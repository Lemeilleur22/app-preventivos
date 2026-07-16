from PIL import Image as PILImage, ImageOps
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage, LongTable, HRFlowable
import streamlit as st
import pandas as pd
from supabase import create_client
import re
import unicodedata
from dotenv import load_dotenv
from datetime import datetime, timezone, date, timedelta
import random
import uuid
import base64
import streamlit.components.v1 as components
import smtplib
import matplotlib.pyplot as plt
from email.message import EmailMessage
from zoneinfo import ZoneInfo
import plotly.express as px
from io import BytesIO
from html import escape
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
st.set_page_config(layout="wide")

st.markdown("""
<style>
div.stButton > button {
    background: linear-gradient(135deg, #e30613, #ff2b2b);
    color: white;
    padding: 16px 45px;
    border-radius: 12px;
    font-size: 18px;
    font-weight: 600;
    border: none;
    box-shadow: 0 10px 25px rgba(227,6,19,0.35);
}
</style>
""", unsafe_allow_html=True)

# CARGAR VARIABLES
load_dotenv()

SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("Faltan variables del entorno de Supabase")
    st.stop()

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

MESES_MTTR = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY",
    6: "JUN", 7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT",
    11: "NOV", 12: "DIC"
}

MESES_CORTOS = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY",
    6: "JUN", 7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT",
    11: "NOV", 12: "DIC"
}

TIPOS_FALLA_MTTR = [
    "MECANICA",
    "ELECTRICA",
    "ESTRUCTURAL",
    "HIDRAULICA",
    "MANTENIMIENTO MAYOR"
]


@st.cache_data(ttl=60)
def cargar_tecnicos_mttr():
    data = supabase.table("tecnicos").select("*").execute().data or []

    tecnicos = []
    for t in data:
        activo = t.get("activo", True)
        if activo is True or str(activo).lower() in ["true", "1", "si", "sí"]:
            tecnicos.append(t)
    return sorted(tecnicos, key=lambda x: x.get("nombre", ""))


@st.cache_data(ttl=60)
def cargar_equipos_mttr():
    return [{"id": None, "nombre": "SIN DEFINIR"}]


def comprimir_imagen(uploaded_file, max_px=900, calidad=50, max_kb=180):
    uploaded_file.seek(0)
    imagen = PILImage.open(uploaded_file)
    imagen = ImageOps.exif_transpose(imagen)
    imagen.thumbnail((max_px, max_px))

    if imagen.mode in ("RGBA", "P"):
        fondo = PILImage.new("RGB", imagen.size, (255, 255, 255))
        fondo.paste(imagen)
        imagen = fondo
    else:
        imagen = imagen.convert("RGB")

    salida = BytesIO()

    for calidad_actual in range(calidad, 24, -5):
        salida = BytesIO()
        imagen.save(
            salida,
            format="JPEG",
            quality=calidad_actual,
            optimize=True,
            progressive=True
        )
        if salida.tell() <= max_kb * 1024:
            break
    salida.seek(0)
    return salida.getvalue()


def subir_evidencia_comprimida(upload_file, prefijo, numero_ot):
    nombre_archivo = f"{prefijo}_{numero_ot}_{uuid.uuid4().hex}.jpg"
    imagen_bytes = comprimir_imagen(upload_file)
    supabase.storage.from_("evidencias").upload(
        nombre_archivo,
        imagen_bytes,
        file_options={"content-type": "image/jpeg"}
    )
    return supabase.storage.from_("evidencias").get_public_url(nombre_archivo)


def subir_evidencia_mttr(upload_file, fecha_actividad):
    nombre_archivo = (
        f"mttr_mtbf/{fecha_actividad.year}/"
        f"{fecha_actividad.month:02d}/"
        f"MTTR_{uuid.uuid4().hex}.jpg"
    )
    imagen_bytes = comprimir_imagen(
        upload_file,
        max_px=900,
        calidad=50,
        max_kb=180
    )
    supabase.storage.from_("evidencias").upload(
        nombre_archivo,
        imagen_bytes,
        file_options={"content-type": "image/jpeg"}
    )
    url = supabase.storage.from_("evidencias").get_public_url(nombre_archivo)
    return url, nombre_archivo


@st.cache_data(ttl=300)
def cargar_historico_refacciones():
    data = (
        supabase
        .table("historico_correctivos")
        .select("*")
        .execute()
    )
    return pd.DataFrame(data.data)


@st.cache_data(ttl=30)
def cargar_ultimos_mttr(responsable_email):
    data = (
        supabase
        .table("registros_mttr_mtbf")
        .select("*")
        .eq("responsable_email", responsable_email)
        .order("created_at", desc=True)
        .limit(3)
        .execute()
        .data or []
    )
    return pd.DataFrame(data)


def vista_supervisor_mttr_mtbf():
    st.subheader("Captura MTTR/MTBF")

    usuario = st.session_state.get("user")
    email_usuario = getattr(usuario, "email", "")

    responsable_db = supabase.table("tecnicos") \
        .select("*") \
        .eq("email", email_usuario) \
        .limit(1) \
        .execute().data
    if responsable_db:
        responsable = responsable_db[0]
        responsable_id = str(responsable.get("id"))
        responsable_nombre = responsable.get("nombre", email_usuario)
    else:
        responsable_id = None
        responsable_nombre = email_usuario

    tecnicos = cargar_tecnicos_mttr()
    equipos = cargar_equipos_mttr()

    if not tecnicos:
        st.error("No hay técnicos cargados en Supabase")
        st.stop()
    if not equipos:
        st.error("No hay equipos cargados en Supabase")
        st.stop()
    hoy = date.today()

    if st.session_state.pop("mttr_guardado_ok", True):
        st.success("Registro guardado correctamente.")
    st.markdown("### Ultimos 3 registros cargados")

    df_ultimos = cargar_ultimos_mttr(email_usuario)

    if df_ultimos.empty:
        st.info("Aun no hay registros capturados por este supervisor.")
    else:
        columnas_ultimos = [
            "fecha_registro",
            "tecnico_nombre",
            "tipo_de_falla",
            "equipo_nombre",
            "tiempo_reparacion_min",
            "descripcion"
        ]

        columnas_ultimos = [
            c for c in columnas_ultimos if c in df_ultimos.columns]

        st.dataframe(
            df_ultimos[columnas_ultimos],
            use_container_width=True,
            hide_index=True
        )

    hoy = date.today()
    col_mes, col_anio, col_dia = st.columns(3)

    with col_anio:
        anios = list(range(2024, hoy.year + 1))
        anio = st.selectbox("Año", anios, index=len(
            anios) - 1, key="mttr_anio")

    with col_mes:
        meses_disponibles = list(range(1, 13))
        if anio == hoy.year:
            meses_disponibles = list(range(1, hoy.month + 1))

        mes_num = st.selectbox(
            "Mes a registrar",
            meses_disponibles,
            index=len(meses_disponibles) - 1,
            format_func=lambda m: MESES_MTTR[m],
            key="mttr_mes"
        )

        primer_dia = date(anio, mes_num, 1)

        if mes_num == 12:
            ultimo_dia_mes = date(anio + 1, 1, 1) - timedelta(days=1)
        else:
            ultimo_dia_mes = date(anio, mes_num + 1, 1) - timedelta(days=1)

        ultimo_dia = min(ultimo_dia_mes, hoy)
        fecha_default = hoy if anio == hoy.year and mes_num == hoy.month else primer_dia

        with col_dia:
            fecha_actividad = st.date_input(
                "Dia",
                value=fecha_default,
                min_value=primer_dia,
                max_value=ultimo_dia,
                key=f"mttr_fecha_{anio}_{mes_num}"
            )

        with st.form("form_mttr_mtbf", clear_on_submit=True):
            descripcion = st.text_area("Descripcion", height=130)

            tecnico_sel = st.selectbox(
                "Tecnico que realizo la actividad",
                tecnicos,
                format_func=lambda t: t.get("nombre", "Sin nombre")
            )
            tipo_falla = st.selectbox("Tipo de falla", TIPOS_FALLA_MTTR)
            equipos_sel = st.selectbox(
                "Equipo",
                equipos,
                format_func=lambda e: f"{e.get('nombre', 'Sin nombre')} - {e.get('sede', 'Sin sede')}"
            )

            col_h, col_m = st.columns(2)

            with col_h:
                horas = st.number_input("Horas", min_value=0, step=1)

            with col_m:
                minutos = st.number_input(
                    "Minutos", min_value=0, max_value=59, step=1)

            evidencia_mttr = st.file_uploader(
                "Evidencia fotografica opcional",
                type=["jpg", "jpeg", "png"]
            )

            st.info(f"Responsable: {responsable_nombre}")
            guardar = st.form_submit_button(
                "Guardar Registro", width="stretch")

    if guardar:
        if not descripcion.strip():
            st.error("La descripcion es obligatoria")
            st.stop()
        tiempo_total_segundos = int(horas * 3600 + minutos * 60)
        if tiempo_total_segundos <= 0:
            st.error("El tiempo de reparacion debe ser mayor a 0.")
            st.stop()

        evidencia_url = None
        evidencia_path = None

        if evidencia_mttr is not None:
            evidencia_url, evidencia_path = subir_evidencia_mttr(
                evidencia_mttr,
                fecha_actividad
            )

        registro = {
            "anio": int(anio),
            "mes_num": int(mes_num),
            "mes": MESES_MTTR[mes_num],
            "fecha_registro": fecha_actividad.isoformat(),
            "descripcion": descripcion.strip(),
            "tecnico_id": str(tecnico_sel.get("id")),
            "tecnico_nombre": tecnico_sel.get("nombre"),
            "tipo_de_falla": tipo_falla,
            "equipo_id": str(equipos_sel.get("id")),
            "equipo_nombre": equipos_sel.get("nombre"),
            "sede": equipos_sel.get("sede"),
            "tiempo_horas": int(horas),
            "tiempo_minutos": int(minutos),
            "tiempo_segundos": 0,
            "tiempo_total_segundos": tiempo_total_segundos,
            "tiempo_reparacion_min": round(tiempo_total_segundos / 60, 2),
            "responsable_id": responsable_id,
            "responsable_nombre": responsable_nombre,
            "responsable_email": email_usuario,
            "evidencia_url": evidencia_url,
            "evidencia_path": evidencia_path,
            "evidencia_nombre": evidencia_mttr.name if evidencia_mttr else None,
            "user_id": getattr(usuario, "id", None)
        }
        supabase.table("registros_mttr_mtbf").insert(registro).execute()
        cargar_ultimos_mttr.clear()
        st.session_state["mttr_guardado_ok"] = True
        st.rerun()


@st.cache_data(ttl=60)
def cargar_preventivos_dashboard():
    preventivos = supabase.table("preventivos").select("*").execute().data
    tecnicos_db = supabase.table("tecnicos") \
        .select("id,nombre,area,activo") \
        .execute().data

    df_prev = pd.DataFrame(preventivos or [])

    if df_prev.empty:
        return df_prev

    for col in [
        "numero_ot", "descripcion", "sede", "estatus", "tecnico_id",
        "schedfinish", "fecha_inicio", "fecha_cierre", "created_at"
    ]:
        if col not in df_prev.columns:
            df_prev[col] = None

    df_tec = pd.DataFrame(tecnicos_db or [])

    if df_tec.empty:
        df_tec = pd.DataFrame(
            columns=["tecnico_id", "tecnico", "area", "activo"])
    else:
        df_tec = df_tec.rename(columns={
            "id": "tecnico_id",
            "nombre": "tecnico"
        })

    df = df_prev.merge(
        df_tec[["tecnico_id", "tecnico", "area", "activo"]],
        on="tecnico_id",
        how="left"
    )
    df["tecnico"] = df["tecnico"].fillna("Sin asignar")
    df["area"] = df["area"].fillna("Sin area")
    df["sede"] = df["sede"].fillna("Sin sede")
    df["estatus"] = df["estatus"].fillna("").astype(str).str.upper()

    df["schedfinish_dt"] = pd.to_datetime(df["schedfinish"], errors="coerce")

    df["created_at_dt"] = pd.to_datetime(
        df["created_at"],
        errors="coerce",
        utc=True
    ).dt.tz_convert("America/Mexico_City")

    df["fecha_inicio_dt"] = pd.to_datetime(
        df["fecha_inicio"],
        errors="coerce",
        utc=True
    ).dt.tz_convert("America/Mexico_City")

    df["fecha_cierre_dt"] = pd.to_datetime(
        df["fecha_cierre"],
        errors="coerce",
        utc=True
    ).dt.tz_convert("America/Mexico_City")

    df["ANIO"] = df["created_at_dt"].dt.year
    df["MES_NUM"] = df["created_at_dt"].dt.month
    df["MES"] = df["MES_NUM"].map(MESES_CORTOS).fillna("SIN FECHA")

    hoy = date.today()

    df["vencida"] = (
        df["estatus"].isin(["PENDIENTE", "EN PROCESO"])
        & df["schedfinish_dt"].notna()
        & (df["schedfinish_dt"].dt.date < hoy)
    )

    df["duracion_horas"] = (
        df["fecha_cierre_dt"] - df["fecha_inicio_dt"]
    ).dt.total_seconds() / 3600

    return df


def enviar_correo(destinatarios, asunto, html):
    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = st.secrets["SMTP_FROM"]
    msg["To"] = ", ".join(destinatarios)
    msg.set_content("Informe semanal de preventivos.")
    msg.add_alternative(html, subtype="html")

    with smtplib.SMTP(st.secrets["SMTP_HOST"], int(st.secrets["SMTP_PORT"])) as smtp:
        smtp.starttls()
        smtp.login(st.secrets["SMTP_USER"], st.secrets["SMTP_PASSWORD"])
        smtp.send_message(msg)


def enviar_correo_con_pdf(destinatarios, asunto, cuerpo, pdf_bytes, nombre_archivo):
    msg = EmailMessage()
    msg["Subject"] = asunto
    msg["From"] = st.secrets["SMTP_FROM"]
    msg["To"] = ", ".join(destinatarios)
    msg.set_content(cuerpo)

    msg.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=nombre_archivo
    )

    with smtplib.SMTP(st.secrets["SMTP_HOST"], int(st.secrets["SMTP_PORT"])) as smtp:
        smtp.starttls()
        smtp.login(st.secrets["SMTP_USER"], st.secrets["SMTP_PASSWORD"])
        smtp.send_message(msg)


def generar_pdf_reporte_preventivos(df_reporte, firma_nombre):
    buffer = BytesIO()
    df = df_reporte.copy()

    if "schedfinish_dt" not in df.columns:
        df["schedfinish_dt"] = pd.to_datetime(
            df["schedfinish"], errors="coerce")
    hoy_ts = pd.Timestamp(date.today())
    df["dias_para_vencer"] = (
        df["schedfinish_dt"].dt.normalize() - hoy_ts
    ).dt.days

    abiertos = df["estatus"].isin(["PENDIENTE", "EN PROCESO"])

    df_pendientes = df[df["estatus"] == "PENDIENTE"].copy()
    df_vencidas = df[df["vencida"] == True].copy()
    df_proximas = df[
        abiertos
        & df["dias_para_vencer"].between(0, 3, inclusive="both")
    ].copy()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=28,
        leftMargin=28,
        topMargin=26,
        bottomMargin=28
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleClean",
        parent=styles["Title"],
        fontName="Times-Bold",
        fontSize=20,
        alignment=1,
        spaceAfter=10
    )
    h2 = ParagraphStyle(
        "H2Clean",
        parent=styles["Heading2"],
        fontName="Times-Bold",
        fontSize=13,
        spaceBefore=12,
        spaceAfter=6
    )
    small = ParagraphStyle(
        "SmallTable",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=10,
        leading=12
    )
    normal = ParagraphStyle(
        "SmallTable",
        parent=styles["Normal"],
        fontName="Times-Roman",
        fontSize=10,
        leading=12
    )

    elementos = []

    total = len(df_reporte)
    pendientes = len(df_reporte[df_reporte["estatus"] == "PENDIENTE"])
    en_proceso = len(df_reporte[df_reporte["estatus"] == "EN PROCESO"])
    realizadas = len(df_reporte[df_reporte["estatus"] == "REALIZADO"])
    vencidas = len(df_reporte[df_reporte["vencida"] == True])
    cumplimiento = round((realizadas / total) * 100, 1) if total else 0

    elementos.append(
        Paragraph("Reporte de Preventivos (Viernes)", title_style))
    elementos.append(HRFlowable(
        width="100%", thickness=0.8, color=colors.black))
    elementos.append(Spacer(1, 12))

    resumen = [
        ["Total OTs", "Pendientes", "En Proceso", "Vencidas", "Cumplimiento"],
        [total, pendientes, en_proceso, vencidas, f"{cumplimiento}%"]
    ]

    tabla_resumen = Table(resumen, colWidths=[130, 130, 130, 130, 130])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e30613")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Times-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabla_resumen)
    elementos.append(Spacer(1, 16))

    estado = df_reporte["estatus"].value_counts()
    labels = estado.index.tolist()
    values = estado.values.tolist()
    colores_estado = {
        "REALIZADO": "#636efa",
        "PENDIENTE": "#EF553B",
        "EN PROCESO": "#00CC96"
    }

    fig, ax = plt.subplots(figsize=(5.8, 3.4))

    ax.pie(
        values,
        labels=labels,
        colors=[colores_estado.get(x, "#ab63fa") for x in labels],
        autopct="%1.1f%%",
        startangle=90,
        wedgeprops={"width": 0.42},
        textprops={"fontsize": 9}
    )
    ax.set_title("Estado actual de preventivos")
    ax.axis("equal")

    img_estado = BytesIO()
    plt.savefig(img_estado, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    img_estado.seek(0)
    elementos.append(RLImage(img_estado, width=330, height=205))
    elementos.append(Spacer(1, 10))

    columnas = ["numero_ot", "descripcion", "tecnico",
                "area", "sede"]

    def agregar_tabla(titulo, df_tabla, rojo=False):
        elementos.append(Paragraph(titulo, h2))

        if df_tabla.empty:
            elementos.append(Paragraph("Sin registros", normal))
            elementos.append(Spacer(1, 8))
            return
        df_mostrar = df_tabla[[
            c for c in columnas if c in df_tabla.columns]].copy()

        data = [[Paragraph(escape(str(c)), small) for c in df_mostrar.columns]]

        for _, row in df_mostrar.iterrows():
            data.append([
                Paragraph(escape(str(row.get(c, "")))[:300], small)
                for c in df_mostrar.columns
            ])

        tabla = LongTable(
            data,
            repeatRows=1,
            colWidths=[68, 250, 86, 78, 78, 70, 66]
        )

        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f2f2")),
            ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.7),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]

        if rojo:
            estilo.append(("BACKGROUND", (0, 1), (-1, -1),
                          colors.HexColor("#fde2e2")))
            estilo.append(("TEXTCOLOR", (0, 1), (-1, -1),
                          colors.HexColor("#990000")))

        tabla.setStyle(TableStyle(estilo))
        elementos.append(tabla)
        elementos.append(Spacer(1, 10))

    agregar_tabla("OTs pendientes", df_pendientes)
    agregar_tabla("OTs vencidas", df_vencidas, rojo=True)
    agregar_tabla("OTs proximas a vencer", df_proximas)

    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph("Atentamente", normal))
    elementos.append(
        Paragraph(f"<b>{escape(firma_nombre)}</b>", normal))

    doc.build(elementos)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def enviar_reporte_semanal_supervisores():
    hoy = date.today()
    inicio = hoy - timedelta(days=hoy.weekday())
    fin = inicio + timedelta(days=6)

    inicio_anterior = inicio - timedelta(days=7)
    fin_anterior = inicio - timedelta(days=1)

    tz_mx = ZoneInfo("America/Mexico_City")

    inicio_anterior_carga = datetime.combine(
        inicio_anterior,
        datetime.min.time(),
        tzinfo=tz_mx
    ).astimezone(timezone.utc)

    fin_anterior_carga = datetime.combine(
        fin_anterior + timedelta(days=1),
        datetime.min.time(),
        tzinfo=tz_mx
    ).astimezone(timezone.utc)

    inicio_carga = datetime.combine(
        inicio,
        datetime.min.time(),
        tzinfo=tz_mx
    ).astimezone(timezone.utc)

    fin_carga = datetime.combine(
        fin + timedelta(days=1),
        datetime.min.time(),
        tzinfo=tz_mx
    ).astimezone(timezone.utc)

    preventivos = supabase.table("preventivos") \
        .select("*") \
        .gte("created_at", inicio_carga.isoformat()) \
        .lt("created_at", fin_carga.isoformat()) \
        .execute().data

    preventivos_anterior = supabase.table("preventivos") \
        .select("*") \
        .gte("created_at", inicio_anterior_carga.isoformat()) \
        .lt("created_at", fin_anterior_carga.isoformat()) \
        .in_("estatus", ["PENDIENTE", "EN PROCESO"]) \
        .execute().data

    if not preventivos:
        st.warning("No hay preventivos programado para esta semana")
        return

    tecnicos_db = supabase.table("tecnicos") \
        .select("id,nombre,area") \
        .execute().data

    df_prev = pd.DataFrame(preventivos)
    df_anterior = pd.DataFrame(preventivos_anterior)

    pendientes_anteriores_por_tecnico = {}

    if not df_anterior.empty:
        pendientes_anteriores_por_tecnico = (
            df_anterior.groupby("tecnico_id")
            .size()
            .to_dict()
        )

    df_tec = pd.DataFrame(tecnicos_db).rename(columns={
        "id": "tecnico_id",
        "nombre": "tecnico"
    })

    df = df_prev.merge(df_tec, on="tecnico_id", how="left")
    df = df.sort_values(["tecnico", "schedfinish", "numero_ot"])

    total = len(df)
    realizados = len(df[df["estatus"] == "REALIZADO"])
    pendientes = len(df[df["estatus"].isin(["PENDIENTE", "EN PROCESO"])])

    bloques_tecnicos = ""

    for tecnico, grupo in df.groupby("tecnico"):
        area = grupo["area"].iloc[0] if "area" in grupo.columns else ""
        total_tecnico = len(grupo)
        tecnico_id = grupo["tecnico_id"].iloc[0]
        pendientes_tecnico = pendientes_anteriores_por_tecnico.get(
            tecnico_id, 0)

        filas = ""

        for _, row in grupo.iterrows():

            filas += f"""
            <tr>
                <td>{row.get("numero_ot", "")}</td>
                <td>{row.get("descripcion", "")}</td>
                <td>{row.get("sede", "")}</td>
            </tr>
            """
        bloques_tecnicos += f"""
        <div class="tecnico-card">
            <h3>{tecnico}</h3>
            <p class="meta">
                <b>Area:</b> {area} |
                <b>Total:</b> {total_tecnico} |
                <b>Pendientes semana anterior:</b> {pendientes_tecnico}
            </p>
            
            <table>
                <thead>
                    <tr>
                        <th>OT</th>
                        <th>Descripcion</th>
                        <th>Sede</th>
                    </tr>
                </thead>
                <tbody>
                    {filas}
                </tbody>
            </table>
        </div>
        """

    html = f"""
    <html>
    <body>
        <style>
            body {{
                font-family: Arial, sans-serif;
                color: #222;
            }}
            h2 {{
                color: #c00000;
                margin-bottom: 5px;
            }}
            .resumen {{
                background: #f4f4f4;
                padding: 12px;
                border-left: 5px solid #c00000;
                margin: 15px 0 25px 0;
            }}
            .tecnico-card {{
                margin-bottom: 28px;
                padding-bottom: 18px;
                border-bottom: 2px solid #eeeeee;
            }}
            .tecnico-card h3 {{
                color: #111;
                margin-bottom: 5px;
            }}
            .meta {{
                font-size: 13px;
                margin-top: 0;
                margin-bottom: 10px;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                font-size: 12px;
            }}
            th {{
                background: #c00000;
                color: white;
                padding: 8px;
                border: 1px solid #ddd;
                text-align: left;
            }}
            td {{
                padding: 7px;
                border: 1px solid #ddd;
                vertical-align: top;
            }}
            tr:nth-child(even) {{
                background: #f8f8f8;
            }}
        </style>
        
        <h2>Reporte semanal de preventivos</h2>
        <p><b>Semana:</b> {inicio.strftime('%d/%m/%Y')} al {fin.strftime('%d/%m/%Y')}</p>

        <div class="resumen">
            <b>Total de preventivos:</b> {total}<br>
            <b>Realizados:</b> {realizados}<br>
            <b>Pendientes / En proceso:</b> {pendientes}
        </div>
        
        {bloques_tecnicos}
    </body>
    </html>
    """
    supervisores = dict(st.secrets["SUPERVISORES"])
    destinatarios = list(supervisores.values())

    enviar_correo(
        destinatarios,
        f"Reporte semanal de preventivos - {inicio.strftime('%d/%m/%Y')}",
        html
    )
    st.success("Reporte enviado correctamente a los supervisores.")


def mostrar_detalle_ejecucion(fecha_inicio, fecha_cierre):
    if not fecha_inicio or pd.isna(fecha_inicio):
        st.warning("Sin fecha de inicio")
        return

    tz = "America/Mexico_City"

    try:
        inicio = pd.to_datetime(fecha_inicio, utc=True).tz_convert(tz)
    except Exception as e:
        st.error(f"Error fecha inicio: {e}")
        return

    inicio_fmt = inicio.strftime("%d/%m/%Y %I:%M %p")
    fin_fmt = "En proceso"
    duracion_fmt = "Calculando..."

    if fecha_cierre and pd.notna(fecha_cierre):
        try:
            cierre = pd.to_datetime(fecha_cierre, utc=True).tz_convert(tz)
            fin_fmt = cierre.strftime("%d/%m/%Y %I:%M %p")
            duracion = cierre - inicio

            if duracion.total_seconds() >= 0:
                total_segundos = int(duracion.total_seconds())
                dias = total_segundos // 86400
                horas = (total_segundos % 86400) // 3600
                minutos = (total_segundos % 3600) // 60

                duracion_fmt = f"{dias} días, {horas}horas, {minutos}min"
            else:
                duracion_fmt = "Error: Cierre antes de inicio"
        except Exception as e:
            fin_fmt = "Error leyendo fecha"
            duracion_fmt = str(e)

    html = f"""
    <div style="
        background:white;
        padding:25px;
        border-radius:18px;
        box-shadow:0 6px 20px rgba(0,0,0,0.08);
        margin:20px 0;
    ">
        <h2>Detalle de ejecución</h2>
        <p><b>Inicio:</b> {inicio_fmt}</p>
        <p><b>Fin:</b> {fin_fmt}</p>
        <p><b>Duración:</b> {duracion_fmt}</p>
    </div>
    """
    components.html(html, height=220)


def img_to_base64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


def guardar_anotaciones_si_hay_cambios(ot_sel, texto_db):
    key = f"anotaciones_{ot_sel}"

    texto_actual = st.session_state.get(key, "")
    texto_actual = str(texto_actual or "").strip()
    texto_db = str(texto_db or "").strip()

    if texto_actual != texto_db:
        supabase.table("preventivos").update({
            "anotaciones": texto_actual
        }).eq("numero_ot", ot_sel).execute()


def obtener_rol_app(email):
    email = email.strip().lower()
    if email == "ethanmijail22@gmail.com":
        return "admin"
    usuario_app = supabase.table("usuarios_app") \
        .select("rol,activo") \
        .eq("email", email) \
        .eq("activo", True) \
        .limit(1) \
        .execute().data
    if usuario_app:
        return usuario_app[0]["rol"]
    return None


ESTATUS_REFACCIONES = ["PENDIENTE", "SOLICITADA", "FINALIZADA"]


def obtener_correos_refacciones():
    correos = st.secrets.get("CORREOS_REFACCIONES", [])

    if isinstance(correos, str):
        return [c.strip() for c in correos.split(",") if c.strip()]
    return list(correos)


def enviar_correo_refaccion(solicitud):
    destinatarios = obtener_correos_refacciones()

    if not destinatarios:
        raise ValueError("No hay correos en CORREOS_REFACCIONES.")

    msg = EmailMessage()
    msg["Subject"] = f"Solicitud de refacción - {solicitud['refaccion']}"
    msg["From"] = st.secrets["SMTP_USER"]
    msg["To"] = ", ".join(destinatarios)

    msg.set_content(f"""
Nueva solicitud de refacción

Supervisor: {solicitud['supervisor_email']}

Refacción a solicitar: {solicitud['refaccion']}
Número de modelo: {solicitud['numero_modelo']}
Piezas: {solicitud['piezas']}
Marca: {solicitud['marca']}
                    
""")

    with smtplib.SMTP(st.secrets["SMTP_HOST"], int(st.secrets.get("SMTP_PORT", 587))) as smtp:
        smtp.starttls()
        smtp.login(st.secrets["SMTP_USER"], st.secrets["SMTP_PASSWORD"])
        smtp.send_message(msg)


@st.cache_data(ttl=30)
def cargar_solicitudes_refacciones_supervisor(supervisor_email):
    data = (
        supabase.table("solicitudes_refacciones")
        .select("*")
        .eq("supervisor_email", supervisor_email)
        .eq("oculta_supervisor", False)
        .order("created_at", desc=True)
        .execute()
        .data or []
    )
    return pd.DataFrame(data)


@st.cache_data(ttl=30)
def cargar_solicitudes_refacciones_admin():
    data = (
        supabase.table("solicitudes_refacciones")
        .select("*")
        .order("created_at", desc=True)
        .execute()
        .data or []
    )
    return pd.DataFrame(data)


def vista_supervisor_refacciones():
    st.subheader("Solicitud de refacciones")

    supervisor_email = (
        st.session_state.get("email")
        or st.session_state.get("usuario_email")
        or st.session_state.get("usuario")
        or ""
    )

    with st.form("form_solicitud_refaccion", clear_on_submit=True):
        refaccion = st.text_input("Refacción necesitada")
        numero_modelo = st.text_input("Número de modelo")
        piezas = st.number_input("Piezas", min_value=1, step=1)
        marca = st.text_input("Marca")
        equipo = st.text_input("Equipo que necesita la refacción")
        ubicacion = st.text_input("Ubiación/Comedor")

        enviar = st.form_submit_button("Enviar solicitud", width="stretch")

    if enviar:
        campos = [refaccion, numero_modelo, marca, equipo, ubicacion]

        if not all(str(c).strip() for c in campos):
            st.error("Todos loc campos son obligatorios")
        else:
            solicitud = {
                "supervisor_email": supervisor_email,
                "refaccion": refaccion.strip(),
                "numero_modelo": numero_modelo.strip(),
                "piezas": int(piezas),
                "marca": marca.strip(),
                "equipo": equipo.strip(),
                "ubicacion": ubicacion.strip(),
                "estatus": "PENDIENTE"
            }

            supabase.table("solicitudes_refacciones").insert(
                solicitud).execute()
            enviar_correo_refaccion(solicitud)

            cargar_solicitudes_refacciones_supervisor.clear()
            cargar_solicitudes_refacciones_admin.clear()

            st.success("Solicitud enviada correctamente.")
            st.rerun()

    st.markdown("---")
    st.subheader("Solicitudes pendientes")

    df = cargar_solicitudes_refacciones_supervisor(supervisor_email)

    if df.empty:
        st.info("No tienes solicitudes pendientes.")
        return

    for _, row in df.iterrows():
        with st.container(border=True):
            st.write(f"**Refacción:** {row['refaccion']}")
            st.write(
                f"**Modelo:** {row['numero_modelo']} | **Marca:** {row['marca']} | **Piezas:** {row['piezas']}")
            st.write(
                f"**Equipo:** {row['equipo']} | **Ubicación:** {row['ubicacion']}")
            st.write(f"**Estatus:** {row['estatus']}")

            if row["estatus"] == "FINALIZADA":
                if st.button("Quitar de pendientes", key=f"ocultar_ref_{row['id']}"):
                    supabase.table("solicitudes_refacciones").update({
                        "oculta_supervisor": True,
                        "oculta_supervisor_at": datetime.now(timezone.utc).isoformat()
                    }).eq("id", row["id"]).execute()

                    cargar_solicitudes_refacciones_supervisor.clear()
                    st.rerun()


def vista_admin_refacciones_solicitudes():
    st.subheader("Solicitudes de refacciones")

    df = cargar_solicitudes_refacciones_admin()

    if df.empty:
        st.info("Aun no hay solicitudes de refacciones")
        return

    for _, row in df.iterrows():
        with st.expander(f"{row['refaccion']} - {row['estatus']} - {row['supervisor_email']}"):
            st.write(f"**Modelo:** {row['numero_modelo']}")
            st.write(f"**Marca:** {row['marca']}")
            st.write(f"**Piezas:** {row['piezas']}")
            st.write(f"**Equipo:** {row['equipo']}")
            st.write(f"**Ubicacion/Comedor:** {row['ubicacion']}")
            st.write(f"**Supervisor:** {row['supervisor_email']}")

            estatus_actual = row.get("estatus", "PENDIENTE")
            nuevo_estatus = st.selectbox(
                "Estatus",
                ESTATUS_REFACCIONES,
                index=ESTATUS_REFACCIONES.index(estatus_actual),
                key=f"estatus_refaccion_{row['id']}"
            )

            comentario = st.text_area(
                "Comentario admin",
                value=row.get("comentario_admin") or "",
                key=f"comentario_refaccion_{row['id']}"
            )

            if st.button("Guardar estatus", key=f"guardar_refaccion_{row['id']}"):
                supabase.table("solicitudes_refacciones").update({
                    "estatus": nuevo_estatus,
                    "comentario_admin": comentario,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }).eq("id", row["id"]).execute()

                if nuevo_estatus == "FINALIZADA":
                    agregar_refaccion_al_historico(row.to_dict())

                cargar_solicitudes_refacciones_admin.clear()
                cargar_solicitudes_refacciones_supervisor.clear()
                cargar_historico_refacciones.clear()
                obtener_inventario_refacciones.clear()
                st.success("Estatus actualizado")
                st.rerun()


def agregar_refaccion_al_historico(solicitud):
    fecha_creacion = pd.to_datetime(
        solicitud.get("created_at"),
        errors="coerce"
    )

    if pd.isna(fecha_creacion):
        fecha_inicio = date.today().isoformat()
    else:
        fecha_inicio = fecha_creacion.date().isoformat()

    registro = {
        "solicitud_id": solicitud["id"],
        "numero_cm": f"REF-{str(solicitud['id'])[:8].upper()}",
        "refaccion_solicitada": solicitud["refaccion"],
        "cantidad": int(solicitud["piezas"]),
        "fecha_inicio": fecha_inicio,
        "estatus": "REALIZADO",
        "fecha_cierre": date.today().isoformat(),
        "craeted_by": solicitud.get("supervisor_email"),
        "rol": "supervisor",
        "supervisor_email": solicitud.get("supervisor_eamil"),
        "numero_modelo": solicitud.get("numero_modelo"),
        "marca": solicitud.get("marca"),
        "equipo": solicitud.get("equipo"),
        "ubicacion": solicitud.get("ubicacion")
    }

    supabase.table("correctivos") \
        .upsert(registro, on_conflict="solicitud_id") \
        .execute()


# ---------------------
# CONTROL DE PANTALLAS
# ---------------------
if "pantalla" not in st.session_state:
    st.session_state.pantalla = "landing"

if st.session_state.pantalla == "landing":

    bg_img = img_to_base64("ilustracion.png")
    logo_img = img_to_base64("acciona_logo.png")

    html = f"""
    <style>
    body {{
        margin: 0;
        font-family: 'Segoe UI', Arial, sans-serif;
    }}

    .hero {{
        width: 100vw;
        height: 100vh;
        background: url("data:image/png;base64,{bg_img}");
        background-size: cover;
        background-position: center;
        display: flex;
    }}

    @media (min-width: 901px) {{
        .hero {{
            position: fixed;
            top: 0;
            left: 0;
            align-items: flex-start;
            padding-top: 120px;
        }}
    }}

    @media (max-width: 900px) {{
        .hero {{
            position: relative;
            align-items: flex-start;
            justify-content: flex-start;
            padding-top: 80px;
            text-align: center;
        }}

        .content {{
            margin-left: 0;
            padding: 0 20px;
        }}

        .title {{
            font-size: clamp(26px, 7vw, 40px);
        }}

        .logo {{
            left: 20px;
            width: 140px;
        }}

    }}

    .hero::after {{
        content: "";
        position: absolute;
        width: 100%;
        height: 100%;
        background: rgba(255,255,255,0.25);
        backdrop-filter: blur(0.5px);
    }}

    .logo {{
        position: absolute;
        top: 30px;
        left: 70px;
        width: 220px;
        z-index: 3;
    }}

    .content {{
        position: relative;
        z-index: 2;
        max-width: 600px;
        margin-left: 80px;
    }}

    .title {{
        font-size: clamp(28px, 5vw, 54px);
        font-weight: 900;
        line-height: 1.05;
        color: #111;
    }}

    .line1 {{
        display: block;
    }}

    .line2 {{
        display: block;
        margin-top: 5px;
    }}

    .subtitle {{
        font-size: clamp(16px, 2vw, 24px);
        margin-top: 15px;
        color: #555;
    }}

    .btn {{
        margin-top: 40px;
        display: inline-block;
        background: linear-gradient(135deg, #e30613, #ff2b2b);
        color: white;
        padding: 16px 45px;
        border-radius: 12px;
        font-size: 18px;
        font-weight: 600;
        text-decoration: none;
        box-shadow: 0 10px 25px rgba(227,6,19,0.35);
        transition: all 0.25s ease;
    }}

    .btn:hover {{
        transform: translateY(-4px) scale(1.02);
        box-shadow: 0 15px 35px rgba(227,6,19,0.5);
    }}

    .card {{
        position: absolute;
        background: rgba(255,255,255,0.75);
        backdrop-filter: blur(6px);
        border-radius: 14px;
        width: 240px;
        height: 130px;
        box-shadow: 0 15px 35px rgba(0,0,0,0.1);
        z-index: 1;
    }}

    .card1 {{
        bottom: 90px;
        left: 120px;
    }}

    .card2 {{
        bottom: 70px;
        right: 140px;
    }}

    /* ANIMACION */
    @keyframes fadeIn {{
        from {{
            opacity: 0;
            transform: translateY(30px);
        }}
        to {{
            opacity: 1;
            transform: translateY(0);
        }}
    }}
    </style>

    <div class="hero">

        <img src="data:image/png;base64,{logo_img}" class="logo">

        <div class="content">

            <div class="title">
                <span class="line1">ACCIONA SERVICIOS URBANOS</span>
                <span class="line2">Y MEDIOAMBIENTALES</span>
            </div>

            <div class="subtitle">
                Sistema de gestión de mantenimiento
            </div>

        </div>

        <div class="card card1"></div>
        <div class="card card2"></div>

    </div>
    """

    valor = components.html(html, height=1000, scrolling=False)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown("<div style='margin-top:-80px'></div>",
                    unsafe_allow_html=True)

        if st.button("Acceder", key="btn_logic_real", use_container_width=True):
            st.session_state.pantalla = "login"
            st.rerun()

    st.stop()

if st.session_state.pantalla == "login":
    colum1, colum2, colum3 = st.columns([1, 2, 1])

    with colum2:
        st.subheader("👷‍♂️ Login")

        usuario_input = st.text_input("Usuario")
        password = st.text_input("Contraseña", type="password")

        if st.button("Ingresar"):
            try:
                res = supabase.auth.sign_in_with_password({
                    "email": usuario_input,
                    "password": password
                })

                if res.user:
                    tecnico_db = (
                        supabase.table("tecnicos")
                        .select("*")
                        .eq("email", usuario_input)
                        .execute()
                        .data
                    )

                    if tecnico_db:
                        st.session_state.user = res.user
                        st.session_state.tipo = "tecnico"
                        st.session_state.tecnico = tecnico_db[0]
                        st.session_state.pantalla = "app"
                        st.rerun()
                    else:
                        st.error("Usuario no registrado.")
            except Exception as e:
                if "Invalid login credentials" in str(e):
                    st.error("Credenciales incorrectas.")
                else:
                    st.error("Error del sistema. Contacte a soporte.")
                    print(e)

        st.markdown("---")

        # ADMIN
        if st.button("Admin", use_container_width=True):
            st.session_state.pantalla = "admin_login"
            st.rerun()

        if st.button("Volver", use_container_width=True):
            st.session_state.pantalla = "landing"
            st.rerun()

    st.stop()

if st.session_state.pantalla == "admin_login":
    column1, column2, column3 = st.columns([2, 2, 2])

    with column2:
        st.subheader("Login Admin")

        email = st.text_input("Correo")
        password = st.text_input("Contraseña", type="password")

        if st.button("Ingresar Admin"):
            try:
                res = supabase.auth.sign_in_with_password({
                    "email": email,
                    "password": password
                })

                if res.user and res.user.email:
                    st.success("Login correcto")

                    email_login = res.user.email.strip().lower()

                    st.write("Email logeado:", email_login)

                    rol_app = obtener_rol_app(email_login)

                    if rol_app in ["admin", "supervisor"]:
                        st.session_state.user = res.user
                        st.session_state.tipo = rol_app
                        st.session_state.pantalla = "app"
                        st.rerun()

                    else:
                        st.error("No autorizado")

                else:
                    st.error("No se pudo iniciar sesion")

            except Exception as e:
                if "Invalid login credentials" in str(e):
                    st.error("Usuario o contraseña incorrectas.")
                else:
                    st.error("Error del sistema. Por favor contacta con soporte.")
                    print(e)

        if st.button("⬅️ Volver"):
            st.session_state.pantalla = "login"
            st.rerun()

    st.stop()
# ---------------------
# MODO ADMINISTRADOR
# ---------------------
# ----------------
# LOGIN SUPABASE
# ----------------

usuario = st.session_state.get("user")
tipo = st.session_state.get("tipo")

if not st.session_state.get("user"):
    st.session_state.pantalla = "login"
    st.rerun()


@st.cache_data(ttl=60)
def get_tecnicos():
    tecnicos_db = (
        supabase.table("tecnicos")
        .select("*")
        .execute()
        .data
    ) or []

    tecnicos_activos = []

    for tecnico in tecnicos_db:
        activo = tecnico.get("activo", True)

        if (
            activo is True
            or str(activo).strip().lower() in ["true", "1", "si", "sí"]
        ):
            tecnicos_activos.append(tecnico)
    return tecnicos_activos


tecnicos = get_tecnicos()

if tipo == "admin":
    st.success("Modo administrador")
    modo = "Admin"
elif tipo == "supervisor":
    modo = "Supervisor"
else:
    st.info("Modo tecnico")
    modo = "General"

with st.sidebar:
    st.markdown("---")
    st.write("Sesion activa")

    if tipo == "admin":
        st.caption(f"Admin: {usuario.email}")
    elif tipo == "supervisor":
        st.caption(f"Supervisor: {usuario.email}")
    else:
        st.caption(f"Técnico: {st.session_state.tecnico['nombre']}")

    if st.button("🔒 Cerrar sesion", use_container_width=True):
        if modo == "General":
            tecnico_actual = st.session_state.tecnico

            preventivos_abiertos = supabase.table("preventivos") \
                .select("numero_ot, anotaciones") \
                .eq("tecnico_id", tecnico_actual["id"]) \
                .eq("estatus", "EN PROCESO") \
                .execute().data

            for prev in preventivos_abiertos:
                guardar_anotaciones_si_hay_cambios(
                    prev["numero_ot"],
                    prev.get("anotaciones")
                )

        supabase.auth.sign_out()
        st.session_state.clear()
        st.session_state.pantalla = "landing"
        st.rerun()


@st.cache_data(ttl=60)
def cargar_ots_supervisor():
    preventivos = supabase.table(
        "preventivos").select("*").execute().data or []
    tecnicos_db = supabase.table("tecnicos").select(
        "id,nombre,area,activo").execute().data or []

    df = pd.DataFrame(preventivos)

    if df.empty:
        return df

    for col in [
        "numero_ot", "descripcion", "sede", "estatus", "tecnico_id",
        "schedfinish", "created_at", "fecha_inicio", "fecha_cierre",
        "evidencia_inicio_url", "evidencia_fin_url", "anotaciones"
    ]:
        if col not in df.columns:
            df[col] = None

    mapa_nombre = {
        str(t["id"]): t.get("nombre", "Sin nombre")
        for t in tecnicos_db
    }

    mapa_area = {
        str(t["id"]): t.get("area", "Sin area")
        for t in tecnicos_db
    }

    df["tecnico"] = df["tecnico_id"].apply(
        lambda x: mapa_nombre.get(
            str(x), "Sin asignar") if pd.notna(x) else "Sin asignar"
    )
    df["area"] = df["tecnico_id"].apply(
        lambda x: mapa_area.get(
            str(x), "Sin area") if pd.notna(x) else "Sin area"
    )
    df["estatus"] = df["estatus"].fillna("").astype(str).str.upper()
    return df


def vista_supervisor_ots_evidencias():
    st.subheader("OTs por técnico")
    df = cargar_ots_supervisor()

    if df.empty:
        st.info("No hay OTs registradas")
        return
    col1, col2, col3 = st.columns(3)

    with col1:
        tecnicos_opciones = sorted(df["tecnico"].dropna().unique())
        tecnico_sel = st.multiselect(
            "Tecnico",
            ["TODOS"] + tecnicos_opciones,
            default=["TODOS"]
        )
    with col2:
        estatus_opciones = sorted(df["estatus"].dropna().unique())
        estatus_sel = st.multiselect(
            "Estatus",
            ["TODOS"] + estatus_opciones,
            default=["TODOS"]
        )
    with col3:
        areas_opciones = sorted(df["area"].dropna().unique())
        area_sel = st.multiselect(
            "Area",
            ["TODOS"] + areas_opciones,
            default=["TODOS"]
        )
    df_filtrado = df.copy()

    if "TODOS" not in tecnico_sel:
        df_filtrado = df_filtrado[df_filtrado["tecnico"].isin(tecnico_sel)]
    if "TODOS" not in estatus_sel:
        df_filtrado = df_filtrado[df_filtrado["estatus"].isin(estatus_sel)]
    if "TODOS" not in area_sel:
        df_filtrado = df_filtrado[df_filtrado["area"].isin(area_sel)]

    resumen = (
        df_filtrado
        .groupby(["tecnico", "estatus"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    if not resumen.empty:
        columnas_estatus = [c for c in resumen.columns if c != "tecnico"]
        resumen["TOTAL"] = resumen[columnas_estatus].sum(axis=1)
    st.markdown("### Resumen por técnico")
    st.dataframe(resumen, use_container_width=True, hide_index=True)

    columnas_tabla = [
        "numero_ot", "descripcion", "tecnico", "area",
        "sede", "estatus", "schedfinish", "created_at"
    ]
    st.markdown("Detalle de OTs")
    st.dataframe(
        df_filtrado[columnas_tabla],
        use_container_width=True,
        hide_index=True
    )
    st.markdown("---")
    st.markdown("Evidencia de OTs completadas")
    df_realizadas = df_filtrado[df_filtrado["estatus"] == "REALIZADO"].copy()

    if df_realizadas.empty:
        st.info("No hay OTs realizadas con los filtros seleccionados")
        return

    ot_sel = st.selectbox(
        "Selecciona OT realizada",
        df_realizadas["numero_ot"].astype(str).tolist()
    )

    fila = df_realizadas[df_realizadas["numero_ot"].astype(
        str) == ot_sel].iloc[0]

    st.write(f"**Descripción:** {fila.get('descripcion', '')}")
    st.write(f"**Tecnico:** {fila.get('tecnico', '')}")
    st.write(f"**Sede:** {fila.get('sede', '')}")
    st.write(f"**Fecha cierre:** {fila.get('fecha_cierre', '')}")

    if fila.get("anotaciones"):
        st.text_area(
            "Anotaciones",
            value=str(fila.get("anotaciones")),
            height=120,
            disabled=True
        )
    col_ev1, col_ev2 = st.columns(2)

    with col_ev1:
        url_inicio = fila.get("evidencia_inicio_url")
        if url_inicio and isinstance(url_inicio, str) and url_inicio.startswith("http"):
            st.image(url_inicio, caption="Antes de iniciar",
                     use_column_width=True)
        else:
            st.info("Sin evidencia inicial")
    with col_ev2:
        url_fin = fila.get("evidencia_fin_url")
        if url_fin and isinstance(url_fin, str) and url_fin.startswith("http"):
            st.image(url_fin, caption="Despues de finalizar",
                     use_column_width=True)
        else:
            st.info("Sin evidencia final")

# --------------------------------------------
# REGLAS PARA REFACCIONAMIENTO (INICIO)
# --------------------------------------------


def normalizar(texto):
    if pd.isna(texto):
        return ""

    texto = str(texto)

    try:
        texto = texto.encode("latin1").decode("utf-8")
    except:
        pass
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ASCII", "ignore").decode("ASCII")

    texto = texto.upper()
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


reglas = {
    "ASIENTO PARA WC": {
        "modo": "any",
        "palabras": ["ASIENTO", "TAPA DE WC", "TAZA"]
    },

    "CORTINA HAWAIANA": {
        "modo": "any",
        "palabras": ["HAWAIIANAS", "HAWAIANA"]
    },

    "SENSOR": {
        "modo": "all",
        "palabras": ["SENSOR"]
    },

    "CONTACTOR": {
        "modo": "all",
        "palabras": ["CONTACTOR"]
    },

    "RESISTENCIA": {
        "modo": "all",
        "palabras": ["RESISTENCIA"]
    },

    "EMPAQUE": {
        "modo": "all",
        "palabras": ["EMPAQUE"]
    },

    "BATERIAS": {
        "modo": "any",
        "palabras": ["BATERIA", "BATERIAS"]
    },

    "PLAFONES": {
        "modo": "any",
        "palabras": ["PLAFON", "PLAFONES"]
    },

    "BOTONES": {
        "modo": "any",
        "palabras": ["BOTON", "BOTONES"]
    },

    "JABONERAS": {
        "modo": "any",
        "palabras": ["JABONERA", "JABONERAS"]
    },

    "CHAPAS": {
        "modo": "any",
        "palabras": ["CHAPA", "CHAPAS", "BISAGRA", "VISAGRA"]
    },

    "SELENOIDES": {
        "modo": "any",
        "palabras": ["SELENOIDE", "SELENOID", "SOLENOID", "SOLENOIDE"]
    },

    "PINTURA ROJO BERMELLON": {
        "modo": "any",
        "palabras": ["HALLAZGOS HIDRANTE", "HALLAZGOS PREVENTIVO DESCARGA", "MANTTO PREVENTIVO DESCARGA", "MANTTO. PREVENTIVO"]
    },

    "LLAVES PARA LAVAMANOS": {
        "modo": "any",
        "palabras": ["LLAVES PARA LAVAMANO", "LLAVES PARA LAVABO"]
    },

    "KIT DE FILTROS": {
        "modo": "any",
        "palabras": ["KIT DE FILTROS"]
    },

    "TERMISTORES": {
        "modo": "any",
        "palabras": ["TERMISTORES", "TERMISTOR"]
    },

    "SEGURO DE PUERTA": {
        "modo": "any",
        "palabras": ["SEGURO DE PUERTA", "JALADERAS"]
    },

    "EMBOLO PARA FLUXOMETRO": {
        "modo": "any",
        "palabras": ["EMBOLO"]
    },

    "LAMPARAS": {
        "modo": "any",
        "palabras": ["LAMPARA", "FOCO", "FOCOS", "LUMINARIA", "LUMINARIAS"]
    },

    "CIERRAPUERTAS": {
        "modo": "any",
        "palabras": ["CIERRAPUERTAS", "CIERRA PUERTAS"]
    },

    "TERMOSTATOS": {
        "modo": "any",
        "palabras": ["TERMOSTATO", "TERMOSTATOS"]
    },

    "MANGUERAS": {
        "modo": "any",
        "palabras": ["MANGUERA", "MANGUERAS", "COFLEX"]
    },

    "MANOMETROS": {
        "modo": "any",
        "palabras": ["MANOMETRO"]
    },

    "VALVULAS": {
        "modo": "any",
        "palabras": ["VALVULA", "VALVULAS", "ELECTROVALVULA", ]
    },

    "TERMINAL BLOCK": {
        "modo": "any",
        "palabras": ["TERMINAL BLOCK"]
    },

    "BOILER": {
        "modo": "any",
        "palabras": ["BOILER"]
    },

    "CLAVIJAS": {
        "modo": "any",
        "palabras": ["CLAVIJA", "CLAVIJAS"]
    },

    "LLAVES": {
        "modo": "any",
        "palabras": ["LLAVE", "LLAVES"]
    },

    "CABLE": {
        "modo": "any",
        "palabras": ["CABLE", "CABLES"]
    },

    "PINTURA": {
        "modo": "any",
        "palabras": ["PINTURA"]
    },

    "TEE MECANICA SCI": {
        "modo": "any",
        "palabras": ["TEE MECANICA"]
    },

    "FLUJOMETRO": {
        "modo": "any",
        "palabras": ["FLUJOMETRO", "MEDIDOR DE FLUJO"]
    },

    "MODULO DE IGNICION": {
        "modo": "any",
        "palabras": ["MODULO DE IGNICION"]
    },

    "SONDA NUCLEO": {
        "modo": "any",
        "palabras": ["SONDA NUCLEO"]
    },

    "FLUXOMETRO": {
        "modo": "any",
        "palabras": ["FLUXOMETRO"]
    }
}


def clasificar_refaccion(longdesc, description):
    texto = normalizar(f"{longdesc} {description}")

    texto = str(texto).upper()
    texto = " ".join(texto.split())

    for nombre, regla in reglas.items():

        if regla["modo"] == "all":
            if all(p in texto for p in regla["palabras"]):
                return nombre

        elif regla["modo"] == "any":
            if any(p in texto for p in regla["palabras"]):
                return nombre
    return "MANTENIMIENTOS MAYORES"


@st.cache_data(ttl=300)
def obtener_inventario_refacciones():
    df = cargar_historico_refacciones()

    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    df["Refaccion"] = df.apply(
        lambda x: clasificar_refaccion(
            x.get("longdesc")
            or x.get("descripcion_larga")
            or "",
            x.get("description") or ""
        ),
        axis=1
    )
    inventario = (
        df.groupby("Refaccion")
        .size()
        .reset_index(name="Cantidad")
        .sort_values("Cantidad", ascending=False)
    )
    return inventario, df
 # -------------------------------------------
 # REGLAS PARA REFACCIONAMIENTO (FIN)
 # -------------------------------------------


def vista_tecnico_admin(solo_lectura=False):
    st.subheader("Visualizacion modo técnico.")
    col1, col2 = st.columns([1, 2])

    with col1:
        tec_sel_admin = st.selectbox(
            "Visualizar:",
            tecnicos,
            format_func=lambda x: x["nombre"]
        )

    with col2:
        rango = st.date_input(
            "Periodo",
            value=(
                date.today() - timedelta(days=7),
                date.today()
            )
        )

    # Traer preventivos
    prev_admin = supabase.table("preventivos") \
        .select("*") \
        .eq("tecnico_id", tec_sel_admin["id"]) \
        .execute().data

    df_admin_prev = pd.DataFrame(prev_admin)

    try:
        if (
            not df_admin_prev.empty
            and len(rango) == 2
            and "created_at" in df_admin_prev.columns
        ):

            inicio = pd.to_datetime(rango[0]).date()
            fin = pd.to_datetime(rango[1]).date()

            df_admin_prev["created_at"] = pd.to_datetime(
                df_admin_prev["created_at"],
                errors="coerce"
            )

            df_admin_prev = df_admin_prev.dropna(
                subset=["created_at"]
            )
            if not df_admin_prev.empty:
                df_admin_prev["created_at"] = (
                    df_admin_prev["created_at"].dt.date
                )

                df_admin_prev = df_admin_prev[
                    (
                        df_admin_prev["created_at"] >= inicio
                    )
                    &
                    (
                        df_admin_prev["created_at"] <= fin
                    )
                ]

    except Exception as e:
        st.warning(
            f"No fue posible aplicar filtro de fechas: {e}"
        )

    st.write(f"Preventivos de {tec_sel_admin['nombre']}")
    columnas = [
        "numero_ot",
        "descripcion",
        "sede",
        "estatus",
        "created_at"
    ]
    columnas_existentes = [
        c for c in columnas
        if c in df_admin_prev.columns
    ]
    if columnas_existentes:
        st.dataframe(
            df_admin_prev[columnas_existentes],
            use_container_width=True
        )
    else:
        st.warning("No existen columnas para mostrar")

    if not df_admin_prev.empty:
        ot_sel_admin = st.selectbox(
            "Selecciona OT",
            df_admin_prev["numero_ot"],
            key="admin_ot"
        )

        fila = df_admin_prev[df_admin_prev["numero_ot"] == ot_sel_admin]

        if not fila.empty:
            estado_actual = fila.iloc[0]["estatus"]
            evidencia_inicio = fila.iloc[0].get("evidencia_inicio_url")
            evidencia_fin = fila.iloc[0].get("evidencia_fin_url")

            anotaciones = fila.iloc[0].get("anotaciones")
            fecha_inicio = fila.iloc[0].get("fecha_inicio")
            fecha_cierre = fila.iloc[0].get("fecha_cierre")

            st.write(f"Estado actual: **{estado_actual}**")

            if estado_actual == "REALIZADO":
                st.warning("Esta OT ya está cerrada")
            elif estado_actual == "PENDIENTE":
                st.info("Esta OT esta abierta")
            elif estado_actual == "EN PROCESO":
                st.info("Esta OT está en proceso")

            if "confirmar_reasignacion_ot" not in st.session_state:
                st.session_state.confirmar_reasignacion_ot = None
            if "confirmar_quitar_ot" not in st.session_state:
                st.session_state.confirmar_quitar_ot = None

            if not solo_lectura:
                st.markdown("---")
                st.subheader("Gestion de asignacion")

                if estado_actual == "REALIZADO":
                    st.info("Esta OT ya esta cerrada. No se puede reasignar.")
                else:
                    col_reasignar, col_quitar = st.columns(2)

                    with col_reasignar:
                        if st.button("Reasignar OT", key=f"reasignar_{ot_sel_admin}"):
                            st.session_state.confirmar_reasignacion_ot = ot_sel_admin
                            st.session_state.confirmar_quitar_ot = None
                            st.rerun()

                    with col_quitar:
                        if st.button("Quitar asignacion", key=f"quitar_{ot_sel_admin}"):
                            st.session_state.confirmar_quitar_ot = ot_sel_admin
                            st.session_state.confirmar_reasignacion_ot = None
                            st.rerun()

                    if st.session_state.confirmar_reasignacion_ot == ot_sel_admin:
                        tecnicos_destino = [
                            t for t in tecnicos
                            if t["id"] != tec_sel_admin["id"]
                        ]

                        nuevo_tecnico = st.selectbox(
                            "Nuevo tecnico",
                            tecnicos_destino,
                            format_func=lambda x: f"{x['nombre']} - {x['area']}",
                            key=f"nuevo_tecnico_{ot_sel_admin}"
                        )

                        st.warning(
                            f"La OT {ot_sel_admin} dejara de aparecerle a "
                            f"{tec_sel_admin['nombre']} y se asignara a "
                            f"{nuevo_tecnico['nombre']}."
                        )

                        col_confirmar, col_cancelar = st.columns(2)

                        with col_confirmar:
                            if st.button("Confirmar reasignacion", key=f"confirmar_reasignar_{ot_sel_admin}"):
                                supabase.table("preventivos").update({
                                    "tecnico_id": nuevo_tecnico["id"]
                                }).eq("numero_ot", ot_sel_admin).execute()

                                supabase.table("preventivos_no_asignados").delete() \
                                    .eq("numero_ot", ot_sel_admin) \
                                    .execute()

                                supabase.table("preventivos_no_asignados").insert({
                                    "numero_ot": ot_sel_admin,
                                    "descripcion": fila.iloc[0].get("descripcion", ""),
                                    "motivo": f"Quitado de {tec_sel_admin['nombre']} por Admin",
                                    "fecha_carga": datetime.now(timezone.utc).isoformat()
                                }).execute()

                                st.session_state.confirmar_quitar_ot = None
                                st.success("Asignacion quitada correctamente")
                                st.rerun()

                        with col_cancelar:
                            if st.button("Cancelar", key=f"cancelar_quitar_{ot_sel_admin}"):
                                st.session_state.confirmar_quitar_ot = None
                                st.rerun()

        # Mostrar evidencia si existe
            st.markdown("---")
            mostrar_detalle_ejecucion(fecha_inicio, fecha_cierre)

            if solo_lectura:
                anotaciones_admin = str(anotaciones or "")

                st.text_area(
                    "Anotaciones",
                    value=anotaciones_admin,
                    height=150,
                    disabled=True,
                    key=f"supervisor_anotaciones_{ot_sel_admin}"
                )
            else:
                anotaciones_key_admin = f"admin_anotaciones_{ot_sel_admin}"
                if st.session_state.get("ot_admin_anotaciones_cargada") != ot_sel_admin:
                    st.session_state["ot_admin_anotaciones_cargada"] = ot_sel_admin
                    st.session_state[anotaciones_key_admin] = str(
                        anotaciones or "")

                anotaciones_admin = st.text_area(
                    "Anotaciones",
                    height=150,
                    key=anotaciones_key_admin
                )

            col_ev1, col_ev2 = st.columns(2)

            with col_ev1:
                if evidencia_inicio and isinstance(evidencia_inicio, str) and evidencia_inicio.startswith("http"):
                    try:
                        st.image(
                            evidencia_inicio,
                            caption="Antes de iniciar",
                            use_container_width=True
                        )
                    except Exception:
                        st.info("Evidencia inicial invalida")
                else:
                    st.info("Sin evidencia inicial")

                if not solo_lectura:
                    foto_inicio_admin = st.file_uploader(
                        "Subir / reemplazar foto inicial",
                        type=["jpg", "jpeg", "png"],
                        key=f"admin_inicio_{ot_sel_admin}"
                    )

            with col_ev2:
                if evidencia_fin and isinstance(evidencia_fin, str) and evidencia_fin.startswith("http"):
                    try:
                        st.image(
                            evidencia_fin,
                            caption="Despues de finalizar",
                            use_container_width=True
                        )
                    except Exception:
                        st.info("Evidencia final invalida")
                else:
                    st.info("Sin evidencia final")

                if not solo_lectura:
                    foto_fin_admin = st.file_uploader(
                        "Subir/reemplazar foto final",
                        type=["jpg", "jpeg", "png"],
                        key=f"admin_fin_{ot_sel_admin}"
                    )

            if not solo_lectura:
                st.markdown("---")

                col1, col2 = st.columns(2)

                with col1:
                    if st.button("Guardar cambios", key=f"guardar_admin_{ot_sel_admin}"):
                        datos_update = {
                            "anotaciones": anotaciones_admin
                        }
                        if foto_inicio_admin:
                            url_inicio = subir_evidencia_comprimida(
                                foto_inicio_admin,
                                "INICIO_ADMIN",
                                ot_sel_admin
                            )

                            datos_update["evidencia_inicio_url"] = url_inicio

                            if not fecha_inicio:
                                datos_update["fecha_inicio"] = datetime.now(
                                    timezone.utc).isoformat()

                        if foto_fin_admin:
                            url_fin = subir_evidencia_comprimida(
                                foto_fin_admin,
                                "FIN_ADMIN",
                                ot_sel_admin
                            )

                            datos_update["evidencia_fin_url"] = url_fin

                        supabase.table("preventivos").update(datos_update) \
                            .eq("numero_ot", ot_sel_admin) \
                            .execute()

                        st.success("Cambios guardados por Admin")
                        st.rerun()

                with col2:
                    if st.button("Cerrar OT con evidencias", key=f"cerrar_admin_{ot_sel_admin}"):
                        datos_update = {
                            "estatus": "REALIZADO",
                            "fecha_cierre": datetime.now(timezone.utc).isoformat(),
                            "anotaciones": anotaciones_admin
                        }

                        url_inicio = evidencia_inicio
                        url_fin = evidencia_fin

                        if foto_inicio_admin:
                            url_inicio = subir_evidencia_comprimida(
                                foto_inicio_admin,
                                "INICIO_ADMIN",
                                ot_sel_admin
                            )

                        if foto_fin_admin:
                            url_fin = subir_evidencia_comprimida(
                                foto_fin_admin,
                                "FIN_ADMIN",
                                ot_sel_admin
                            )

                        if not url_inicio:
                            st.error(
                                "Debes tener foto inicial antes de cerrar la OT.")
                            st.stop()
                        if not url_fin:
                            st.error(
                                "Debes tener foto final antes de cerrar la OT")
                            st.stop()

                        datos_update["evidencia_inicio_url"] = url_inicio
                        datos_update["evidencia_fin_url"] = url_fin

                        if not fecha_inicio:
                            datos_update["fecha_inicio"] = datetime.now(
                                timezone.utc).isoformat()

                        supabase.table("preventivos").update(datos_update) \
                            .eq("numero_ot", ot_sel_admin) \
                            .execute()

                        st.success("OT cerrada por Admin con evidencias")
                        st.rerun()

        # -------------
        # CERRAR OT
        # -------------
                with col1:
                    if st.button("✅ Forzar cierre"):
                        supabase.table("preventivos").update({
                            "estatus": "REALIZADO",
                            "fecha_cierre": datetime.now(timezone.utc).isoformat()
                        }).eq("numero_ot", ot_sel_admin).execute()

                        st.success("OT cerrada por Admin")
                        st.rerun()

        # -------------
        # REABRIR OT
        # -------------
                with col2:
                    if st.button("🔄️ Reabrir OT"):
                        supabase.table("preventivos").update({
                            "estatus": "PENDIENTE",
                            "fecha_inicio": None,
                            "fecha_cierre": None,
                            "evidencia_inicio_url": None,
                            "evidencia_fin_url": None,
                            "anotaciones": None
                        }).eq("numero_ot", ot_sel_admin).execute()

                        st.success("OT reabierta")
                        st.rerun()
        else:
            st.warning("No hay preventivos para este tecnico")


if modo == "Supervisor":
    tab_sup1, tab_sup2, tab_sup3 = st.tabs([
        "Captura de actividades correctivas",
        "OTs y evidencias",
        "Refacciones"
    ])

    with tab_sup1:
        vista_supervisor_mttr_mtbf()
    with tab_sup2:
        vista_tecnico_admin(solo_lectura=True)
    with tab_sup3:
        vista_supervisor_refacciones()

if modo == "General":
    tecnico_sel = st.session_state.tecnico

if modo == "Admin":

    tab0, tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "HOME",
        "Carga a sistema",
        "Vista técnico",
        "Evidencias",
        "Tecnicos",
        "Refaccionamiento"
    ])

    # ---------------
    # HOME DASHBOARD
    # ---------------
    def obtener_semana_actual():
        hoy = date.today()
        inicio = hoy - timedelta(days=hoy.weekday())
        fin = inicio + timedelta(days=6)
        return inicio, fin

    with tab0:
        st.subheader("Dashboard general")
        df_dash = cargar_preventivos_dashboard()

        if df_dash.empty:
            st.info("No hay preventivos cargados.")
        else:
            st.markdown("""
            <style>
            .dash-card {
                background: white;
                padding: 18px 20px;
                border-radius: 12px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.08);
                border-left: 6px solid #e30613;
                margin-bottom: 12px;
            }
            .dash-title {
                font-size: 14px;
                color: #6b7280;
                font-weight: 600;
            }
            .dash-value {
                font-size: 34px;
                font-weight: 800;
                color: #111827;
            }
            </style>
            """, unsafe_allow_html=True)

            def multiselect_todos(label, opciones, key):
                opciones = sorted([x for x in opciones if pd.notna(x)])
                seleccion = st.multiselect(
                    label,
                    ["TODOS"] + opciones,
                    default=["TODOS"],
                    key=key
                )
                if "TODOS" in seleccion or not seleccion:
                    return opciones
                return seleccion

            semana_inicio, semana_fin = obtener_semana_actual()
            col_f1, col_f2, col_f3, col_f4 = st.columns([1, 1.2, 1, 1])

            with col_f1:
                anios = sorted([int(x)
                               for x in df_dash["ANIO"].dropna().unique()])
                anio_sel = st.selectbox(
                    "Año",
                    ["TODOS"] + anios,
                    key="dash_anio_prev"
                )
            with col_f2:
                meses_sel = multiselect_todos(
                    "Mes",
                    df_dash["MES"].dropna().unique(),
                    "dash_mes_prev"
                )
            with col_f3:
                areas_sel = multiselect_todos(
                    "Area",
                    df_dash["area"].dropna().unique(),
                    "dash_area_prev"
                )
            with col_f4:
                rango_semana = st.date_input(
                    "Semana",
                    value=(semana_inicio, semana_fin),
                    key="dash_rango_semana"
                )

            df_filtrado = df_dash[
                df_dash["MES"].isin(meses_sel)
                & df_dash["area"].isin(areas_sel)
            ].copy()

            if anio_sel != "TODOS":
                df_filtrado = df_filtrado[df_filtrado["ANIO"] == anio_sel]

            if isinstance(rango_semana, tuple) and len(rango_semana) == 2:
                inicio_semana, fin_semana = rango_semana

                df_filtrado = df_filtrado[
                    df_filtrado["created_at_dt"].notna()
                    & (df_filtrado["created_at_dt"].dt.date >= inicio_semana)
                    & (df_filtrado["created_at_dt"].dt.date <= fin_semana)
                ]
            else:
                st.info("Selecciona un rango de fechas para la semana.")

            if df_filtrado.empty:
                st.warning("No hay datos con los filtros seleccionados.")
                df_filtrado = df_dash.iloc[0:0].copy()

            total = len(df_filtrado)
            pendientes = len(
                df_filtrado[df_filtrado["estatus"] == "PENDIENTE"])
            en_proceso = len(
                df_filtrado[df_filtrado["estatus"] == "EN PROCESO"])
            realizados = len(
                df_filtrado[df_filtrado["estatus"] == "REALIZADO"])
            vencidas = len(df_filtrado[df_filtrado["vencida"]])
            cumplimiento = round((realizados / total) * 100, 1) if total else 0

            k1, k2, k3, k4, k5 = st.columns(5)

            metricas = [
                (k1, "Total OTs", total),
                (k2, "Pendientes", pendientes),
                (k3, "En proceso", en_proceso),
                (k4, "Vencidas", vencidas),
                (k5, "Cumplimiento", f"{cumplimiento}%")
            ]

            for col, titulo, valor in metricas:
                with col:
                    st.markdown(f"""
                    <div class="dash-card">
                        <div class="dash-title">{titulo}</div>
                        <div class="dash-value">{valor}</div>
                    </div>        
                    """, unsafe_allow_html=True)

            sub1, sub3 = st.tabs([
                "Resumen",
                "Datos de Técnicos"
            ])

            with sub1:

                ALTURA_GRAFICAS = 360

                estado = (
                    df_filtrado["estatus"]
                    .value_counts()
                    .reset_index()
                )
                estado.columns = ["Estatus", "Total"]

                fig_estado = px.pie(
                    estado,
                    names="Estatus",
                    values="Total",
                    hole=0.60,
                    title="Estado actual de preventivos"
                )
                fig_estado.update_layout(
                    template="plotly_white",
                    height=ALTURA_GRAFICAS,
                    margin=dict(l=10, r=10, t=50, b=10)
                )

                area_resumen = (
                    df_filtrado.groupby("area")
                    .size()
                    .reset_index(name="TOTAL_OT")
                    .sort_values("TOTAL_OT", ascending=False)
                )

                fig_area_resumen = px.pie(
                    area_resumen,
                    names="area",
                    values="TOTAL_OT",
                    hole=0.55,
                    title="Distribucion de OTs por area"
                )
                fig_area_resumen.update_layout(
                    template="plotly_white",
                    height=ALTURA_GRAFICAS,
                    margin=dict(l=10, r=10, t=50, b=10)
                )

                sede_resumen = (
                    df_filtrado["sede"]
                    .value_counts()
                    .head(10)
                    .reset_index()
                )
                sede_resumen.columns = ["Sede", "Total"]

                fig_sede = px.bar(
                    sede_resumen,
                    x="Total",
                    y="Sede",
                    orientation="h",
                    text="Total",
                    title="Top sedes con mas preventivos"
                )
                fig_sede.update_layout(
                    template="plotly_white",
                    height=450,
                    yaxis=dict(autorange="reversed")
                )
                col_estado, col_tabla, col_area = st.columns(
                    [1.45, 0.65, 1.45], gap="small")

                with col_estado:
                    st.plotly_chart(fig_estado, use_container_width=True, config={
                                    "displayModeBar": False})

                with col_tabla:
                    st.dataframe(
                        area_resumen.rename(columns={
                            "area": "Area",
                            "TOTAL_OT": "OTs"
                        }),
                        use_container_width=True,
                        hide_index=True,
                        height=190
                    )

                with col_area:
                    st.plotly_chart(fig_area_resumen, use_container_width=True, config={
                                    "displayModeBar": False})
                st.plotly_chart(fig_sede, use_container_width=True)

            with sub3:
                realizadas_hist = df_dash[
                    df_dash["estatus"] == "REALIZADO"
                ].copy()
                if realizadas_hist.empty:
                    st.info("No hay OTs realizadas.")
                else:
                    historico_tecnico = (
                        realizadas_hist.groupby(["tecnico", "area"])
                        .size()
                        .reset_index(name="OTs realizadas")
                        .sort_values("OTs realizadas", ascending=False)
                    )
                    fig_historico = px.bar(
                        historico_tecnico,
                        x="OTs realizadas",
                        y="tecnico",
                        color="area",
                        orientation="h",
                        text="OTs realizadas",
                        title="Historico de OTs realizadas por tecnico."
                    )
                    fig_historico.update_layout(
                        template="plotly_white",
                        height=max(450, len(historico_tecnico)*32),
                        yaxis=dict(autorange="reversed")
                    )
                    fig_historico.update_traces(
                        textposition="inside"
                    )

                    st.plotly_chart(fig_historico, use_container_width=True, config={
                                    "displayModeBar": False})

                    st.markdown("---")
                    st.subheader("OTs Vencidas")

                    df_vencidas = df_filtrado[
                        df_filtrado["vencida"] == True
                    ].copy()

                    if df_vencidas.empty:
                        st.success(
                            "No hay OTs vencidas con los filtros seleccionados")
                    else:

                        columnas = [
                            "numero_ot",
                            "descripcion",
                            "tecnico",
                            "area",
                            "sede",
                            "schedfinish",
                            "estatus",
                            "vencida"
                        ]

                        columnas = [
                            c for c in columnas
                            if c in df_vencidas.columns
                        ]

                        df_detalle = df_vencidas[columnas].copy()

                        def pintar_vencida(row):
                            return [
                                "background-color: #ffe5e5; color: #8a0000; font-weight: 600"
                                for _ in row
                            ]

                        st.dataframe(
                            df_detalle.style.apply(pintar_vencida, axis=1),
                            use_container_width=True,
                            hide_index=True
                        )

                    st.markdown("---")
                    st.subheader("Enviar reporte PDF")

                    supervisores = dict(st.secrets["SUPERVISORES"])
                    destinatarios = list(supervisores.values())

                    st.caption(
                        f"El reporte se enviara a {len(destinatarios)} supervisores.")

                    firma_nombre = st.text_input(
                        "Firma del reporte",
                        value="Ethan Mijail Chavez Juarez",
                        key="firma_reporte_pdf"
                    )
                    if st.button("Generar PDF y enviar", use_container_width=True):

                        if not destinatarios:
                            st.error("No hay correos registrados.")
                        else:
                            pdf_bytes = generar_pdf_reporte_preventivos(
                                df_filtrado,
                                firma_nombre
                            )
                            nombre_archivo = f"reporte_preventivos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

                            enviar_correo_con_pdf(
                                destinatarios,
                                "Reporte de preventivos",
                                "Adjunto reporte PDF de preventivos",
                                pdf_bytes,
                                nombre_archivo
                            )
                            st.success("Reporte PDF enviado con éxito.")

    # ---------------
    # PRIMERA PESTAÑA
    # ---------------
    with tab1:

        # FILA 1
        col1, col2 = st.columns(2)

        with col1:
            with st.container(border=True):
                st.subheader("Descargar OTs Realizadas")

                rango_cierre = st.date_input(
                    "Rango de fechas de cierre",
                    value=(
                        date.today() - timedelta(days=7),
                        date.today()
                    ),
                    key="rango_descarga_realizadas"
                )

                preventivos_realizados = []

                if isinstance(rango_cierre, tuple) and len(rango_cierre) == 2:
                    inicio, fin = rango_cierre

                    tz_mx = ZoneInfo("America/Mexico_City")

                    inicio_utc = datetime.combine(
                        inicio,
                        datetime.min.time(),
                        tzinfo=tz_mx
                    ).astimezone(timezone.utc)

                    fin_utc = datetime.combine(
                        fin + timedelta(days=1),
                        datetime.min.time(),
                        tzinfo=tz_mx
                    ).astimezone(timezone.utc)

                    preventivos_realizados = (
                        supabase.table("preventivos")
                        .select("*")
                        .eq("estatus", "REALIZADO")
                        .gte("fecha_cierre", inicio_utc.isoformat())
                        .lt("fecha_cierre", fin_utc.isoformat())
                        .execute()
                        .data
                    )
                else:
                    st.info("Selecciona un rango de fechas para descargar.")

                if preventivos_realizados:
                    df_realizados = pd.DataFrame(preventivos_realizados)

                    csv = df_realizados.to_csv(index=False).encode("utf-8-sig")

                    st.download_button(
                        label="Descargar CSV OT realizadas",
                        data=csv,
                        file_name=f"preventivos_realizados_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                else:
                    st.info("No hay preventivos cerrados para este rango.")

        with col2:
            with st.container(border=True):
                st.subheader("Carga Preventivos Maximo")

                archivo = st.file_uploader(
                    "",
                    type=["csv", "xlsx"],
                    key="maximo_excel_v2",
                    label_visibility="collapsed"
                )

                cargar_prev = False
                actualizar_fechas = False
                actualizar_sedes = False

                if archivo:
                    cargar_prev = st.button(
                        "Cargar preventivos",
                        use_container_width=True,
                        key="btn_cargar_v2"
                    )
                    actualizar_fechas = st.button(
                        "Actualizar fechas",
                        use_container_width=True,
                        key="btn_fechas_v2",
                    )
                    actualizar_sedes = st.button(
                        "Actualizar sedes",
                        use_container_width=True,
                        key="btn_sedes_v2"
                    )
        st.markdown("<br>", unsafe_allow_html=True)

        # FILA 2
        col3, col4 = st.columns(2)

        with col3:
            with st.container(border=True):
                st.subheader("Job plans")

                archivo_jp = st.file_uploader(
                    "",
                    type=["xlsx"],
                    key="jobplan_excel",
                    label_visibility="collapsed"
                )

                guardar_jp = False

                if archivo_jp:
                    guardar_jp = st.button(
                        "Guardar Job Plan",
                        use_container_width=True,
                        key="btn_jp"
                    )

        with col4:
            with st.container(border=True):
                st.subheader("Roles")
                archivo_roles = st.file_uploader(
                    "",
                    type=["xlsx"],
                    key="roles_excel",
                    label_visibility="collapsed"
                )

                actualizar_roles = False

                if archivo_roles:
                    actualizar_roles = st.button(
                        "Actualizar turnos",
                        use_container_width=True,
                        key="btn_roles"
                    )

        if archivo and not (cargar_prev or actualizar_fechas or actualizar_sedes):
            st.info("Archivo cargado. Selecciona una opción.")
        if archivo and (cargar_prev or actualizar_fechas or actualizar_sedes):
            try:
                if archivo.name.endswith(".csv"):
                    df = pd.read_csv(archivo)
                else:
                    df = pd.read_excel(archivo)

                df.columns = df.columns.str.lower().str.strip()
                columnas_requeridas = {"wonum", "description", "location"}
                faltantes = columnas_requeridas - set(df.columns)

                if faltantes:
                    st.error(
                        f"El archivo no tiene las siguientes columnas requeridas: {', '.join(faltantes)}")
                    st.stop()
                st.write("Vista previo:", df.head())

                datos_carga = []
                no_asignados = []

                tecnico_cba = next(
                    (
                        t for t in tecnicos
                        if "ISAAC" in str(t.get("nombre", "")).upper()
                        and "ZAMUDIO" in str(t.get("nombre", "")).upper()
                    ),
                    None
                )
                if not tecnico_cba:
                    st.error("No se encontro tecnico CBA")
                    st.stop()

                fontaneros = [
                    t for t in tecnicos
                    if t["area"] == "FONTANERIA"
                    and t["id"] != tecnico_cba["id"]
                ]

                cuota_fontaneria = {
                    t["id"]: 2 for t in fontaneros
                }

                carga_temporal = {
                    t["id"]: 0
                    for t in tecnicos
                    if t["id"] != tecnico_cba["id"]
                }

                tecnico_marin = next(
                    (t for t in tecnicos if t["nombre"] == "Alejandro Marin"),
                    None
                )

                PM_MARIN = "IFMSI006"
                PM_SEGUNDO_TURNO = ["IFMSI214"]
                PM_PRIMER_TURNO = ["IFMSI210"]
                JP_CHECKLIST_BANOS = {"1460123"}
                JP_CHECKLIST_ARRANQUE = {"1460244"}
                JP_TRAMPA_GRASA = {"146994"}

                def limpiar_turno(turno):
                    return str(turno).strip().replace(".0", "")

                for _, row in df.iterrows():

                    location = str(row["location"]).upper()
                    descripcion = str(row["description"]).upper()
                    pmnum = str(row.get("pmnum")).strip().upper()
                    jpnum = str(row.get("jpnum")).strip()
                    fecha_raw = str(row.get("schedfinish", "")).strip()
                    match = re.search(
                        r'([A-Za-z]{3}\s+\d{1,2},\s+\d{4})',
                        fecha_raw
                    )
                    if match:
                        fecha_raw = match.group(1)
                    schedfinish = pd.to_datetime(
                        fecha_raw,
                        format="%b %d, %Y",
                        errors="coerce"
                    )

                    if "NODO" in descripcion:
                        no_asignados.append({
                            "numero_ot": row["wonum"],
                            "descripcion": row["description"],
                            "motivo": "JCI"
                        })
                        continue

                    partes_location = location.split()
                    codigo_sede = location.split()[-1].strip()

                    mapa_sedes = {
                        "ANG": "El Angel",
                        "BA": "Bellas Artes",
                        "CLC": "CLC",
                        "P.P": "Popopiramides",
                        "PIN": "Pintura",
                        "CBA": "CBA"
                    }

                    sede = mapa_sedes.get(codigo_sede, codigo_sede)

                    if ("CBA" in descripcion or codigo_sede == "CBA") and "IFSI" not in location:
                        datos_carga.append({
                            "numero_ot": row["wonum"],
                            "descripcion": row["description"],
                            "jpnum": str(row["jpnum"]).strip() if pd.notna(row.get("jpnum")) else None,
                            "pmnum": pmnum,
                            "estatus": "PENDIENTE",
                            "tecnico_id": tecnico_cba["id"],
                            "sede": sede,
                            "schedfinish": (
                                schedfinish.date().isoformat()
                                if pd.notna(schedfinish)
                                else None
                            )
                        })
                        continue

                    # ----------------
                    # CHECKLIST DE BAÑOS
                    # ----------------
                    if jpnum in JP_CHECKLIST_BANOS:
                        candidatos = [
                            t for t in tecnicos
                            if t["area"] in [
                                "COCINAS",
                                "CONSERVACION",
                                "FONTANERIA"
                            ]
                            and limpiar_turno(
                                t.get("turno_actual")
                            ) in ["1", "2"]
                            and t["id"] != tecnico_cba["id"]
                        ]

                    # --------------------
                    # CHECKLIST DE ARRANQUE
                    # --------------------
                    elif jpnum in JP_CHECKLIST_ARRANQUE:
                        candidatos = [
                            t for t in tecnicos
                            if t["area"] in [
                                "COCINAS",
                                "CONSERVACION"
                            ]
                            and limpiar_turno(t.get("turno_actual")) == "1"
                            and t["id"] != tecnico_cba["id"]
                        ]
                    # -----------------
                    # TRAMPAS DE GRASA
                    # -----------------
                    elif jpnum in JP_TRAMPA_GRASA:
                        candidatos = [
                            t for t in tecnicos
                            if t["area"] in [
                                "COCINAS",
                                "CONSERVACION"
                            ]
                            and limpiar_turno(
                                t.get("turno_actual")
                            ) in ["2", "3"]
                            and t["id"] != tecnico_cba["id"]
                        ]

                    # ---------------
                    # SCI
                    # ---------------
                    elif "IFSI" in location:

                        # CCTV MARIN
                        if pmnum == PM_MARIN:

                            if tecnico_marin:
                                datos_carga.append({
                                    "numero_ot": row["wonum"],
                                    "descripcion": row["description"],
                                    "jpnum": str(row["jpnum"]).strip() if pd.notna(row.get("jpnum")) else None,
                                    "pmnum": pmnum,
                                    "estatus": "PENDIENTE",
                                    "tecnico_id": tecnico_marin["id"],
                                    "sede": sede,
                                    "schedfinish": (
                                        schedfinish.date().isoformat()
                                        if pd.notna(schedfinish)
                                        else None
                                    )
                                })
                            continue

                        # SOLO SEGUNDO TURNO
                        elif (
                            pmnum in PM_SEGUNDO_TURNO
                            or "C-IFSI-CBA" in location
                            or "C-IFSI-CLC" in location
                        ):
                            candidatos = [
                                t for t in tecnicos
                                if t["area"] == "SCI"
                                and limpiar_turno(t.get("turno_actual")) == "2"
                                and t["id"] != tecnico_cba["id"]
                            ]

                        # SOLO PRIMER TURNO
                        elif pmnum in PM_PRIMER_TURNO:
                            candidatos = [
                                t for t in tecnicos
                                if t["area"] == "SCI"
                                and limpiar_turno(t.get("turno_actual")).strip() == "1"
                                and t["id"] != tecnico_cba["id"]
                            ]

                        # ----------
                        # RESTO SCI
                        # ----------
                        else:
                            candidatos = [
                                t for t in tecnicos
                                if t["area"] == "SCI"
                                and t["id"] != tecnico_cba["id"]
                            ]

                    # ----------------
                    # IFCO
                    # ----------------
                    elif "IFCO" in location:
                        candidatos = [
                            t for t in tecnicos
                            if t["area"] in ["COCINAS", "CONSERVACION"]
                            and t["id"] != tecnico_cba["id"]
                        ]

                    # ----------------
                    # IFFA
                    # ----------------
                    elif "IFFA" in location:

                        fontaneros_pendientes = [
                            t for t in fontaneros
                            if cuota_fontaneria[t["id"]] > 0
                        ]

                        if fontaneros_pendientes:
                            candidatos = fontaneros_pendientes
                        else:
                            candidatos = [
                                t for t in tecnicos
                                if t["area"] in ["COCINAS", "CONSERVACION"]
                                and t["id"] != tecnico_cba["id"]
                            ]

                    # ---------------
                    # NO ASIGNAR
                    # ---------------
                    elif "IFVI" in location or "IFDE" in location:
                        no_asignados.append({
                            "numero_ot": row["wonum"],
                            "descripcion": row["description"],
                            "motivo": "PROVEEDOR"
                        })
                        continue

                    else:
                        no_asignados.append({
                            "numero_ot": row["wonum"],
                            "descripcion": row["description"],
                            "motivo": "Sin regla de asignacion"
                        })
                        continue

                    if not candidatos:
                        no_asignados.append({
                            "numero_ot": row["wonum"],
                            "descripcion": row["description"],
                            "motivo": "No hay tecnico disponible"
                        })
                        continue

                    min_carga = min(carga_temporal[t["id"]]
                                    for t in candidatos)

                    tecnicos_disponibles = [
                        t for t in candidatos
                        if carga_temporal[t["id"]] == min_carga
                    ]

                    tecnico_asignado = random.choice(tecnicos_disponibles)

                    # Actualizar contador
                    carga_temporal[tecnico_asignado["id"]] += 1

                    if tecnico_asignado["area"] == "FONTANERIA":
                        cuota_fontaneria[tecnico_asignado["id"]] -= 1

                    jpnum_val = row.get("jpnum")

                    if pd.notna(jpnum_val):
                        jpnum_val = str(jpnum_val).strip()
                    else:
                        jpnum_val = None

                    datos_carga.append({
                        "numero_ot": row["wonum"],
                        "descripcion": row["description"],
                        "jpnum": jpnum_val,
                        "pmnum": pmnum,
                        "estatus": "PENDIENTE",
                        "tecnico_id": tecnico_asignado["id"],
                        "sede": sede,
                        "schedfinish": (
                            schedfinish.date().isoformat()
                            if pd.notna(schedfinish)
                            else None
                        )
                    })

                df_carga = pd.DataFrame(datos_carga)

                if actualizar_sedes:

                    actualizados = 0

                    mapa_sedes = {
                        "ANG": "El Angel",
                        "BA": "Bellas Artes",
                        "CLC": "CLC",
                        "P.P": "Popopiramides",
                        "PIN": "Pintura",
                        "CBA": "CBA"
                    }

                    for _, row in df.iterrows():

                        location = str(row["location"]).upper()
                        partes = location.split()
                        codigo_sede = partes[-1] if partes else None

                        sede = mapa_sedes.get(codigo_sede, codigo_sede)

                        supabase.table("preventivos").update({
                            "sede": sede
                        }).eq("numero_ot", row["wonum"]).execute()

                    actualizados += 1

                    st.success(f"{actualizados} preventivos actualizados")

                if actualizar_fechas:

                    actualizados = 0

                    for _, row in df.iterrows():
                        fecha_raw = str(row.get("schedfinish", "")).strip()
                        match = re.search(
                            r'([A-Za-z]{3}\s+\d{1,2},\s+\d{4})',
                            fecha_raw
                        )
                        if match:
                            fecha_raw = match.group(1)
                        schedfinish = pd.to_datetime(
                            fecha_raw,
                            format="%b %d, %Y",
                            errors="coerce"
                        )

                        supabase.table("preventivos").update({
                            "schedfinish": (
                                schedfinish.date().isoformat()
                                if pd.notna(schedfinish)
                                else None
                            )
                        }).eq(
                            "numero_ot",
                            row["wonum"]
                        ).execute()

                        actualizados += 1
                        st.success(
                            f"{actualizados} OTS actualizadas"
                        )

                if cargar_prev:

                    existentes = supabase.table("preventivos") \
                        .select("numero_ot") \
                        .execute().data
                    ots_existentes = {x["numero_ot"] for x in existentes}

                    nuevos = [
                        row for row in df_carga.to_dict(orient="records")
                        if row["numero_ot"] not in ots_existentes
                    ]

                    if nuevos:
                        supabase.table("preventivos").insert(nuevos).execute()

                    supabase.table("preventivos_no_asignados") \
                        .delete() \
                        .neq("numero_ot", "") \
                        .execute()

                    if no_asignados:
                        for item in no_asignados:
                            item["fecha_carga"] = datetime.now(
                                timezone.utc).isoformat()

                        supabase.table("preventivos_no_asignados") \
                            .insert(no_asignados) \
                            .execute()

                    st.success(
                        f"{len(nuevos)} preventivos cargados | "
                        f"{len(no_asignados)} no asignados"
                    )
            except Exception as e:
                st.error(
                    "Ocurrio un error al procesar el archivo de carga de preventivos.")
                st.exception(e)
                st.stop()

        # ---------------
        # JOB PLANS
        # ---------------

        if archivo_jp:
            excel_file = pd.ExcelFile(archivo_jp)
            hojas = excel_file.sheet_names

            todas_tareas = []
            for hoja in hojas:
                df_excel = pd.read_excel(
                    archivo_jp,
                    sheet_name=hoja,
                    header=None
                )

                # -----------------
                # OBTENER JPNUM
                # -----------------
                import re

                jpnum = None

                for _, row in df_excel.iterrows():
                    for valor in row:
                        texto = str(valor)

                        match = re.search(r'JPNUM[:\s]*([0-9]+)', texto)

                        if match:
                            jpnum = match.group(1)
                            break

                    if jpnum:
                        break

                if not jpnum:
                    st.error("No se pudo detectar el JPNUM")
                    st.stop()

                st.success(f"JP Detectado: {jpnum}")

                # -------------------------
                # ENCONTRAR INICIO DE TASKS
                # -------------------------
                inicio_tasks = None

                for i, row in df_excel.iterrows():

                    if "TASK ID" in str(row[0]).upper():
                        inicio_tasks = i + 1
                        break

                tareas = []

                if inicio_tasks:

                    i = inicio_tasks

                    while i < len(df_excel):
                        secuencia = df_excel.iloc[i, 0]
                        descripcion = df_excel.iloc[i, 1]

                        if pd.isna(secuencia) and pd.isna(descripcion):
                            i += 1
                            continue

                        if pd.notna(secuencia):

                            try:
                                secuencia = int(secuencia)
                                texto_completo = str(descripcion).strip()

                                # Leer filas inferiores
                                j = i + 1

                                while j < len(df_excel):
                                    siguiente_sec = df_excel.iloc[j, 0]
                                    texto_extra = df_excel.iloc[j, 1]

                                    # Si encuentra nueva secuencia -> termina
                                    if pd.notna(siguiente_sec):
                                        break

                                    # Agregar texto adicional
                                    if pd.notna(texto_extra):
                                        texto_completo += "\n" + \
                                            str(texto_extra).strip()

                                    j += 1

                                todas_tareas.append({
                                    "jpnum": str(jpnum),
                                    "secuencia": secuencia,
                                    "tarea": texto_completo
                                })

                                i = j

                            except Exception as e:
                                st.warning(f"Fila ignorada: {e}")
                                i += 1
                        else:
                            i += 1

            # -----------------
            # FINAL DATAFRAME
            # -----------------
            df_tareas = pd.DataFrame(todas_tareas)
            st.dataframe(df_tareas)

            # ------------
            # GUARDAR
            # ------------

            if guardar_jp:
                insertados = 0

                for jp in df_tareas["jpnum"].unique():
                    existente = supabase.table("jobplan_tareas") \
                        .select("jpnum") \
                        .eq("jpnum", jp) \
                        .execute().data

                    if existente:
                        continue

                    datos_jp = df_tareas[
                        df_tareas["jpnum"] == jp
                    ]

                    supabase.table("jobplan_tareas").insert(
                        datos_jp.to_dict(orient="records")
                    ).execute()

                    insertados += 1

                st.success(f"{len(df_tareas)} tareas guardadas.")

        if archivo_roles:
            from openpyxl import load_workbook

            wb = load_workbook(archivo_roles, data_only=True)
            ws = wb.active

            fila_header = None

            for fila in ws.iter_rows(min_row=1, max_row=30):
                valores = [
                    str(cell.value).strip().upper()
                    for cell in fila
                    if cell.value is not None
                ]

                if "NOMBRE" in valores and "TURNO" in valores:
                    fila_header = fila[0].row
                    break

            st.write("Fila detectada:", fila_header)

            if fila_header is None:
                st.error("No se encontró encabezado")
                st.stop()

            datos = list(ws.values)

            encabezados = [
                str(x).strip().upper() if x is not None else ""
                for x in datos[fila_header - 1]
            ]

            df_roles = pd.DataFrame(
                datos[fila_header:],
                columns=encabezados
            )

            df_roles = df_roles[
                (df_roles["NOMBRE"].notna()) &
                (df_roles["TURNO"].notna())
            ]

            st.dataframe(df_roles[["NOMBRE", "TURNO"]], hide_index=True)

            def limpiar_texto(txt):
                txt = str(txt)
                txt = txt.replace("\xa0", " ")
                txt = str(txt).upper().strip()
                txt = re.sub(r"\s+", " ", txt)
                return txt

            if actualizar_roles:

                actualizados = 0
                no_encontrados = []

                tecnicos_db = supabase.table("tecnicos").select(
                    "id,nombre_nomina"
                ).execute().data

                mapa_tecnicos = {}

                for t in tecnicos_db:
                    nombre_limpio = limpiar_texto(t["nombre_nomina"])
                    mapa_tecnicos[nombre_limpio] = t["id"]

                for _, row in df_roles.iterrows():
                    nombre_excel = limpiar_texto(row["NOMBRE"])
                    turno = str(row["TURNO"]).strip()

                    if nombre_excel in ["", "NAN", "NONE"]:
                        continue

                    tecnico_id = mapa_tecnicos.get(nombre_excel)

                    if tecnico_id:
                        supabase.table("tecnicos").update({
                            "turno_actual": turno
                        }).eq("id", tecnico_id).execute()

                        actualizados += 1
                    else:
                        no_encontrados.append(nombre_excel)

                st.success(f"{actualizados} técnicos actualizados")

                if no_encontrados:
                    st.warning("No encontrados:")
                    st.write(no_encontrados)
                st.cache_data.clear()
                st.success("Roles actualizados")
                st.rerun()
        with st.container(border=True):
            st.subheader("Reporte semanal a supervisores")
            if st.button("Enviar reporte semanal", use_container_width=True):
                enviar_reporte_semanal_supervisores()

    # ----------------
    # SEGUNDA PESTAÑA
    # ----------------
    with tab2:
        vista_tecnico_admin(solo_lectura=False)

    # ---------------
    # TERCERA PESTAÑA
    # ---------------
    with tab3:
        st.subheader("Auditoría")

        termino = st.text_input(
            "Buscar por número OT, descripción o sede",
            placeholder="Ej: A86354164, baños, comedor, CLC..."
        )

        if termino:
            termino = termino.strip()

            resultados = supabase.table("preventivos") \
                .select("*") \
                .or_(
                    f"numero_ot.ilike.%{termino}%,"
                    f"descripcion.ilike.%{termino}%,"
                    f"sede.ilike.%{termino}%"
            ) \
                .execute().data

            df_resultados = pd.DataFrame(resultados)

            if df_resultados.empty:
                st.warning("No se encontraron preventivos.")
            else:
                st.dataframe(
                    df_resultados[[
                        "numero_ot",
                        "descripcion",
                        "sede",
                        "estatus"
                    ]],
                    use_container_width=True,
                    hide_index=True
                )

                ot_sel_audit = st.selectbox(
                    "Selecciona OT para auditoría",
                    df_resultados["numero_ot"],
                    key="auditoria_ot"
                )

                fila = df_resultados[
                    df_resultados["numero_ot"] == ot_sel_audit
                ]

                if not fila.empty:
                    estado_actual = fila.iloc[0]["estatus"]
                    evidencia_inicio = fila.iloc[0].get("evidencia_inicio_url")
                    evidencia_fin = fila.iloc[0].get("evidencia_fin_url")
                    anotaciones = fila.iloc[0].get("anotaciones")
                    fecha_inicio = fila.iloc[0].get("fecha_inicio")
                    fecha_cierre = fila.iloc[0].get("fecha_cierre")
                    tecnico_id = fila.iloc[0].get("tecnico_id")

                    tecnico_nombre = "Sin asignar"

                    if tecnico_id:
                        tecnico = supabase.table("tecnicos") \
                            .select("nombre") \
                            .eq("id", tecnico_id) \
                            .execute().data

                        if tecnico:
                            tecnico_nombre = tecnico[0]["nombre"]

                    st.markdown("---")

                    st.write(f"**OT:** {ot_sel_audit}")
                    st.write(f"**Estado:** {estado_actual}")
                    st.write(f"**Técnico asignado:** {tecnico_nombre}")

                    mostrar_detalle_ejecucion(
                        fecha_inicio,
                        fecha_cierre
                    )

                    if anotaciones:
                        st.text_area(
                            "Anotaciones del técnico",
                            value=anotaciones,
                            height=180,
                            disabled=True
                        )
                    else:
                        st.info("Sin anotaciones")

                    col1, col2 = st.columns(2)

                    with col1:
                        if evidencia_inicio and isinstance(evidencia_inicio, str) and evidencia_inicio.startswith("http"):
                            try:
                                st.image(
                                    evidencia_inicio,
                                    caption="Evidencia inicial",
                                    use_container_width=True
                                )
                            except Exception:
                                st.info("Evidencia incial invalida")
                        else:
                            st.info("Sin evidencia inicial")

                    with col2:
                        if evidencia_fin and isinstance(evidencia_fin, str) and evidencia_fin.startswith("http"):
                            try:
                                st.image(
                                    evidencia_fin,
                                    caption="Evidencia final",
                                    use_container_width=True
                                )
                            except Exception:
                                st.info("Evidencia final invalida")
                        else:
                            st.info("Sin evidencia final")

        st.markdown("---")
        st.subheader("Preventivos no asignados")

        no_asignados_db = supabase.table("preventivos_no_asignados") \
            .select("*") \
            .order("fecha_carga", desc=True) \
            .execute().data

        if no_asignados_db:
            df_no = pd.DataFrame(no_asignados_db)

            st.dataframe(
                df_no[["numero_ot", "descripcion", "motivo", "fecha_carga"]],
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No hay preventivos sin asignar")

    with tab4:
        st.subheader("Gestion de personal")

        tecnicos_todos = (
            supabase.table("tecnicos")
            .select("*")
            .execute()
            .data
        )

        for tecnico in tecnicos_todos:

            col1, col2 = st.columns([3, 1])

            with col1:
                estado = "Activo" if tecnico["activo"] else "Inactivo"
                turno_raw = tecnico.get("turno_actual")

                if turno_raw is None or turno_raw == "":
                    turno = "Sin turno"
                else:
                    try:
                        turno = str(int(float(turno_raw)))
                    except:
                        turno = str(turno_raw)

                st.write(
                    f"{tecnico['nombre']} - {tecnico['area']} - Turno {turno} - {estado}")

            with col2:
                nuevo_estado = st.toggle(
                    "Activo",
                    value=tecnico["activo"],
                    key=f"tec_{tecnico['id']}"
                )

                if nuevo_estado != tecnico["activo"]:
                    supabase.table("tecnicos").update({
                        "activo": nuevo_estado
                    }).eq("id", tecnico["id"]).execute()

                    st.cache_data.clear()
                    st.rerun()

    with tab5:
        st.subheader("Historico de Refacciones")

        inventario, historial = obtener_inventario_refacciones()
        st.dataframe(
            inventario,
            use_container_width=True,
            hide_index=True
        )

        fig = px.bar(
            inventario,
            x="Cantidad",
            y="Refaccion",
            orientation="h",
            text="Cantidad",
            title="Historico de refacciones"
        )
        fig.update_layout(
            template="plotly_white",
            height=max(500, len(inventario) * 28),
            yaxis=dict(autorange="reversed")
        )
        st.plotly_chart(fig, use_container_width=True)

        vista_admin_refacciones_solicitudes()

# --------------------
# PUBLICO GENERAL
# --------------------
if "confirmar_cierre_ot" not in st.session_state:
    st.session_state.confirmar_cierre_ot = None

if modo == "General":
    st.subheader("👷‍♂️ Vista de colaboradores")

    preventivos = supabase.table("preventivos") \
        .select("*") \
        .eq("tecnico_id", tecnico_sel["id"]) \
        .in_("estatus", ["PENDIENTE", "EN PROCESO"]) \
        .execute().data

    df_prev = pd.DataFrame(preventivos)

    if not df_prev.empty:
        st.dataframe(
            df_prev[[
                "numero_ot",
                "descripcion",
                "sede",
                "estatus"
            ]],
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("No tienes OTs pendientes. Gracias por tu trabajo. 🙂")

    if not df_prev.empty:
        ot_sel = st.selectbox(
            "Selecciona OT",
            df_prev["numero_ot"]
        )
        fila_ot = df_prev[df_prev["numero_ot"] == ot_sel]

        jpnum_ot = fila_ot.iloc[0]["jpnum"]
        jpnum_ot = str(jpnum_ot).strip()

        tareas_jp = supabase.table("jobplan_tareas") \
            .select("*") \
            .eq("jpnum", str(jpnum_ot)) \
            .order("secuencia") \
            .execute().data

        with st.expander("Detalles de tarea"):
            if tareas_jp:
                for tarea in tareas_jp:
                    st.markdown(
                        f"**{tarea['secuencia']}** - {tarea['tarea']}"
                    )
            else:
                st.warning(
                    "No hay Job Plan cargado. Comuniquese con su planner.")

        estado_ot = fila_ot.iloc[0]["estatus"]

        # --------------
        # OT PENDIENTE
        # --------------
        if estado_ot == "PENDIENTE":
            st.warning("Tarea aun no iniciada")

            foto_inicio = st.file_uploader(
                "Sube foto iniciando actividad",
                type=["jpg", "jpeg", "png"],
                key=f"inicio_{ot_sel}"
            )

            if st.button("Iniciar tarea"):

                if not foto_inicio:
                    st.error("Debes subir una foto inicial de la actividad.")
                else:
                    nombre_inicio = f"INICIO_{ot_sel}_{uuid.uuid4().hex}.jpg"

                    supabase.storage.from_("evidencias").upload(
                        nombre_inicio,
                        foto_inicio.getvalue()
                    )

                    url_inicio = supabase.storage.from_(
                        "evidencias"
                    ).get_public_url(nombre_inicio)

                    supabase.table("preventivos").update({
                        "estatus": "EN PROCESO",
                        "evidencia_inicio_url": url_inicio,
                        "fecha_inicio": datetime.now(timezone.utc).isoformat()
                    }).eq("numero_ot", ot_sel).execute()

                    st.success("Tarea iniciada ✅")
                    st.rerun()

        # ---------------
        # OT EN PROCESO
        # ---------------
        elif estado_ot == "EN PROCESO":
            st.info("Tarea en proceso")

            evidencia_inicio = fila_ot.iloc[0].get("evidencia_inicio_url")
            evidencia_fin = fila_ot.iloc[0].get("evidencia_fin_url")
            anotacion_actual = str(fila_ot.iloc[0].get("anotaciones") or "")

            key_anot = f"anotaciones_{ot_sel}"
            if key_anot not in st.session_state:
                st.session_state[key_anot] = str(anotacion_actual or "")

            # --------------
            # FOTO INICIAL
            # --------------
            if evidencia_inicio and isinstance(evidencia_inicio, str) and evidencia_inicio.startswith("http"):
                try:
                    st.image(
                        evidencia_inicio,
                        caption="Foto inicial",
                        width=200
                    )

                    if st.button("Reemplazar foto inicial"):
                        guardar_anotaciones_si_hay_cambios(
                            ot_sel,
                            anotacion_actual
                        )

                        supabase.table("preventivos").update({
                            "evidencia_inicio_url": None
                        }).eq("numero_ot", ot_sel).execute()

                        st.session_state[key_anot] = str(
                            st.session_state.get(key_anot, ""))
                        st.success("Ahora sube la nueva foto incial")
                        st.rerun()

                except Exception:
                    st.warning(
                        "La evidencia inicial esta corrupta o es invalida")
                    evidencia_inicio = None

            else:
                st.warning("La evidencia inicial es inválida o inexistente")

                nueva_inicio = st.file_uploader(
                    "Sube foto inicial",
                    type=["jpg", "png", "jpeg"],
                    key=f"reemplazo_inicio_{ot_sel}"
                )

                if st.button("Guardar nueva foto incial"):
                    if not nueva_inicio:
                        st.error("Debes subir una evidencia")
                    else:
                        nombre_inicio = f"INICIO_{ot_sel}_{uuid.uuid4().hex}.jpg"

                        supabase.storage.from_("evidencias").upload(
                            nombre_inicio,
                            nueva_inicio.getvalue()
                        )

                        url_inicio = supabase.storage.from_(
                            "evidencias"
                        ).get_public_url(nombre_inicio)

                        supabase.table("preventivos").update({
                            "evidencia_inicio_url": url_inicio
                        }).eq("numero_ot", ot_sel).execute()

                        st.session_state[key_anot] = str(
                            st.session_state.get(key_anot, ""))
                        st.success("Foto inicial reemplazada")
                        st.rerun()

            # --------------
            # ANOTACIONES
            # --------------
            key_anot = f"anotaciones_{ot_sel}"

            anotaciones = st.text_area(
                "Anotaciones de la actividad",
                key=key_anot,
                height=150,
                placeholder="Ejemplo: Se detecto fuga en baño X, se reemplazó empaque"
            )

            if st.button("Guardar anotaciones"):
                guardar_anotaciones_si_hay_cambios(
                    ot_sel,
                    anotacion_actual
                )
                st.success("Anotaciones guardadas correctamente")
                st.rerun()

            # ----------------
            # EVIDENCIA FINAL
            # ----------------
            if evidencia_fin and isinstance(evidencia_fin, str) and evidencia_fin.startswith("http"):
                try:
                    st.image(
                        evidencia_fin,
                        caption="Foto final actual",
                        width=200
                    )
                except Exception:
                    st.warning(
                        "La evidencia final esta corrupta o es invalida")
                    evidencia_fin = None

                if evidencia_fin and st.button("Reemplazar foto final"):
                    guardar_anotaciones_si_hay_cambios(
                        ot_sel,
                        anotacion_actual
                    )

                    supabase.table("preventivos").update({
                        "evidencia_fin_url": None
                    }).eq("numero_ot", ot_sel).execute()

                    st.success("Ahora sube la evidencia final")
                    st.rerun()
            else:
                st.warning("Sin evidencia final valida")

            foto_fin = st.file_uploader(
                "Sube evidencia de fin de tarea",
                type=["jpg", "jpeg", "png"],
                key=f"fin{ot_sel}"
            )

            # --------------------
            # CIERRE DE EVIDENCIAS
            # --------------------

            if st.button("Finalizar tarea"):
                st.session_state.confirmar_cierre_ot = ot_sel
                st.rerun()

            if st.session_state.confirmar_cierre_ot == ot_sel:
                st.warning(
                    f"Estás a punto de cerrar el preventivo:\n\n"
                    f"**{fila_ot.iloc[0]['descripcion']}**\n\n"
                    f"¿Estás seguro?"
                )

                col1, col2 = st.columns(2)

                with col1:
                    if st.button("Sí, finalizar"):
                        guardar_anotaciones_si_hay_cambios(
                            ot_sel,
                            anotacion_actual
                        )

                        evidencia_inicio_actual = supabase.table("preventivos") \
                            .select("evidencia_inicio_url,evidencia_fin_url") \
                            .eq("numero_ot", ot_sel) \
                            .execute().data[0]

                        if not evidencia_inicio_actual["evidencia_inicio_url"]:
                            st.error("Debes contar con foto inicial")
                            st.stop()

                        url_fin = evidencia_inicio_actual["evidencia_fin_url"]

                        if foto_fin:
                            nombre_fin = f"FIN_{ot_sel}_{uuid.uuid4().hex}.jpg"

                            supabase.storage.from_("evidencias").upload(
                                nombre_fin,
                                foto_fin.getvalue()
                            )

                            url_fin = supabase.storage.from_(
                                "evidencias"
                            ).get_public_url(nombre_fin)

                        if not url_fin:
                            st.error("Debes subir evidencia final")
                            st.stop()

                        supabase.table("preventivos").update({
                            "estatus": "REALIZADO",
                            "fecha_cierre": datetime.now(timezone.utc).isoformat(),
                            "evidencia_fin_url": url_fin,
                            "anotaciones": anotaciones
                        }).eq("numero_ot", ot_sel).execute()

                        st.session_state.confirmar_cierre_ot = None
                        st.success("OT Cerrada ✅")
                        st.rerun()

                with col2:
                    if st.button("Regresar"):
                        st.session_state.confirmar_cierre_ot = None
                        st.rerun()
